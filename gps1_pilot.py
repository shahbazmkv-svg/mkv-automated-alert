"""
gps1_pilot_v3.py — MKV CAR RENTAL Pilot GPS Daily Report (REST API v3 edition)

Drop-in replacement for gps1_pilot.py. Uses the documented Pilot GPS REST API v3
instead of the reverse-engineered backend scrape.

v2 fixes (from first dry-run 2026-06-17)
────────────────────────────────────────
• DISTANCE_UNIT_DIVISOR corrected: `length` is in km, not meters
  (initial v1 assumed meters → showed total 2.8 km vs expected ~2,800 km)
• `zone` field from /vehicles/status is a {id: name} dict, not a list — extract values
• Plate normalization: replace underscores, split letter-digit runs (B15789mustang),
  title-case all-upper words (CADILLAC → Cadillac)
• `moved_today` now requires length > 0 (filters out idle/sensor-wake "trips")
• Long-window fallback: vehicles with no recent move get a 30-day /distance call
  to recover their actual last-move timestamp (eliminates "—" in idle table on first run)
• Plate column width 35 → 38 chars; smart truncation at word boundary
• Sanity check warns if total daily km is implausibly low (catches future unit drift)
• Known regression: "Engine Blocked: 0" — v3 /vehicles/status sensors for MKV's fleet
  don't expose engine-block state in the same form the legacy current_data.php did.
  Will be addressed via Pilot support ticket.

Key improvements (vs legacy gps1_pilot.py)
──────────────────────────────────────────
• Documented contract (won't break on Pilot UI changes)
• Bearer token auth, cached 48h — 1 login per ~2 days instead of every run
• Batch /vehicles/status call — live state for all vehicles in 1 HTTP request
• Single /vehicles/distance call per vehicle — replaces report_type=6 AND report_type=1

Slack format changes (vs current production output)
───────────────────────────────────────────────────
MSG1 — Snapshot
  • Removed "Moving Now" line
  • Added "Needs Attention: N (X Urgent · Y Watch)" line
  • Added "Outside Dubai: N" counter on the In a Zone line
  • Footer notes SOLD exclusions

MSG2 — Moved Today
  • SOLD-prefixed vehicles excluded fleet-wide
  • Plate names normalized
  • Outside-Dubai zones marked with 🚩
  • Vehicles with trips but 0 km no longer included (filtered out)

MSG3 — Parked, split into two tables
  • "Idle <48h" — recently parked
  • "Long-idle >48h" — review-for-utilization

MSG4 — Speeding, deduped per-vehicle
  • Columns: Vehicle, Peak, Avg, Trips
  • One row per vehicle instead of one per event

Files written
─────────────
gps1_pilot_token.json    — cached bearer token (gitignore this)
gps1_pilot_snapshot.json — odometer + last-move snapshot
MKV_GPS_Report_<DATE>.xlsx — Excel report (5 tabs)

Env vars (same as gps1_pilot.py)
────────────────────────────────
PILOT_USER, PILOT_PASS, SLACK_TOKEN, SLACK_CHANNEL
Optional: PILOT_HOST (default https://pilot-gps.ru)

Usage
─────
    python gps1_pilot_v3.py
    python gps1_pilot_v3.py --date 2026-06-15
    python gps1_pilot_v3.py --dry-run      # skip Slack post

Requirements
────────────
    pip install requests openpyxl python-dotenv
"""

import argparse
import io
import json
import os
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path

# Force UTF-8 stdout/stderr on Windows — otherwise `python script.py > file.txt`
# crashes with UnicodeEncodeError on box-drawing chars (─) and emoji (🚗 🚨).
# Two-layer fallback because Python 3.14 has been quirky with stdout reconfigure
# when output is redirected to a file.
if sys.platform == "win32":
    try:
        # Preferred: wrap the underlying binary buffer with a fresh UTF-8 wrapper.
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8",
                                       line_buffering=True, errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8",
                                       line_buffering=True, errors="replace")
    except (AttributeError, io.UnsupportedOperation):
        # Fallback for older Python or unusual stdio: try in-place reconfigure.
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
            sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        except AttributeError:
            pass

import openpyxl
import requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

load_dotenv()

# ════════════════════════════════════════════════════════════════════════
# CONFIG
# ════════════════════════════════════════════════════════════════════════
PILOT_USER    = os.getenv("PILOT_USER", "").strip()
PILOT_PASS    = os.getenv("PILOT_PASS", "").strip()
PILOT_HOST    = os.getenv("PILOT_HOST", "https://pilot-gps.ru").rstrip("/")
SLACK_TOKEN   = os.getenv("SLACK_TOKEN", "")
SLACK_CHANNEL = os.getenv("SLACK_CHANNEL", "C0B6Y6EG85D")

TIMEZONE_NAME = "Asia/Dubai"
DUBAI         = timezone(timedelta(hours=4))

SCRIPT_DIR    = Path(__file__).resolve().parent
TOKEN_FILE    = SCRIPT_DIR / "gps1_pilot_token.json"
SNAPSHOT_FILE = SCRIPT_DIR / "gps1_pilot_snapshot.json"

# Distance API returns length/gps/can in KILOMETERS.
# (Initial assumption was meters; first dry-run on 2026-06-17 showed total 2.8 km
#  vs expected ~2,800 km — clear evidence the field is already km.)
# If a future deployment shows total km < 100 with >10 vehicles moved, flip this back to 1000.
DISTANCE_UNIT_DIVISOR = 1

# Thresholds
SPEED_LIMIT_KMH        = 120   # speeding threshold for MSG4 inclusion
URGENT_SPEED_KMH       = 160   # urgent risk trigger
URGENT_DAILY_KM        = 250   # urgent risk trigger
WATCH_DAILY_KM_MIN     = 150   # watch risk window
URGENT_OFFLINE_HOURS   = 6     # offline >6h = urgent
WATCH_OFFLINE_MINUTES  = 30    # offline >30m = watch (until 6h)
LONG_IDLE_HOURS        = 48    # split threshold for MSG3
WATCH_PARKING_HOURS    = 8     # long parking in unknown zone trigger

# Non-Dubai zone keywords (case-insensitive substring match on zone name).
# We're conservative: only flag when name explicitly contains another emirate.
# Unrecognized zones are reported as "unknown zone", not "outside Dubai".
OUTSIDE_DUBAI_KEYWORDS = [
    "abu dhabi", "sharjah", "ajman", "ras al khaimah", "rak ",
    "umm al quwain", "uaq", "fujairah", "al ain",
]

# SOLD detection — vehicles whose name starts with these tokens are excluded.
SOLD_TOKENS = ("sold ", "sold-", "sold_")

TIMEOUT = 30

# Colors / styles (preserved from gps1_pilot.py to keep Excel identical)
DARK_GREEN = "1A3C2E"
MID_GREEN  = "2E6B4F"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F5F5"
RED        = "CC0000"
AMBER      = "CC7700"
GREY       = "999999"
GREEN      = "2E6B4F"


# ════════════════════════════════════════════════════════════════════════
# EXCEL STYLE HELPERS (verbatim from gps1_pilot.py — preserved on purpose)
# ════════════════════════════════════════════════════════════════════════
def hfont(bold=True, color=WHITE, size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)
def bfont(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def hfill(color=DARK_GREEN):
    return PatternFill("solid", fgColor=color)
def afill():
    return PatternFill("solid", fgColor=LIGHT_GREY)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)
def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = hfont(); c.fill = hfill(); c.alignment = center(); c.border = bdr()
def style_row(ws, row, cols, alt=False):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = bfont(); c.border = bdr(); c.alignment = left()
        if alt: c.fill = afill()
def title_block(ws, title, date_str, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(1, 1, f"MKV CAR RENTAL — {title}")
    t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill = hfill(DARK_GREEN); t.alignment = center()
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    d = ws.cell(2, 1, f"Report Date: {date_str}  |  Timezone: {TIMEZONE_NAME}  |  "
                     f"Generated: {datetime.now(DUBAI).strftime('%Y-%m-%d %H:%M')}")
    d.font = Font(name="Arial", size=9, color=WHITE)
    d.fill = hfill(MID_GREEN); d.alignment = center()
    ws.row_dimensions[2].height = 18
def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ════════════════════════════════════════════════════════════════════════
# TOKEN CACHE
# ════════════════════════════════════════════════════════════════════════
def load_cached_token():
    """Return (token, node_id) if valid cached token exists, else (None, None)."""
    if not TOKEN_FILE.exists():
        return None, None
    try:
        data = json.loads(TOKEN_FILE.read_text(encoding="utf-8"))
        # Refresh 5 minutes before actual expiry for safety
        if data.get("expires_at", 0) > time.time() + 300:
            return data.get("token"), data.get("node_id", 0)
    except (json.JSONDecodeError, OSError):
        pass
    return None, None

def save_token(token, node_id, expires_in):
    """Save bearer token with absolute expiry timestamp."""
    payload = {
        "token": token,
        "node_id": node_id,
        "expires_at": time.time() + int(expires_in or 172800),
    }
    try:
        TOKEN_FILE.write_text(json.dumps(payload), encoding="utf-8")
    except OSError as e:
        print(f"  [warn] could not write {TOKEN_FILE}: {e}")


# ════════════════════════════════════════════════════════════════════════
# SNAPSHOT (for daily-km fallback + long-idle tracking)
# ════════════════════════════════════════════════════════════════════════
def load_snapshot():
    if SNAPSHOT_FILE.exists():
        try:
            return json.loads(SNAPSHOT_FILE.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            pass
    return {"date": "", "odometers": {}, "last_move_ts": {}}

def save_snapshot(vehicles, date_str):
    """
    Persist:
      • odometers — full_mileage per agent_id (for delta-based daily km fallback)
      • last_move_ts — most recent known move-end ts per agent_id (for long-idle days)
    """
    snap = load_snapshot()
    snap["date"] = date_str
    for v in vehicles:
        aid = str(v["agent_id"])
        snap["odometers"][aid] = v["total_km"]
        if v.get("last_move_ts"):
            snap["last_move_ts"][aid] = v["last_move_ts"]
    try:
        SNAPSHOT_FILE.write_text(json.dumps(snap), encoding="utf-8")
    except OSError as e:
        print(f"  [warn] could not write {SNAPSHOT_FILE}: {e}")


# ════════════════════════════════════════════════════════════════════════
# PILOT V3 API CLIENT
# ════════════════════════════════════════════════════════════════════════
class PilotV3Client:
    """Minimal client for Pilot GPS REST API v3 (host: pilot-gps.ru)."""

    def __init__(self, host, username, password):
        self.host = host.rstrip("/")
        self.session = requests.Session()

        token, node_id = load_cached_token()
        if token:
            print(f"  Auth: using cached token (expires in ~"
                  f"{int((json.loads(TOKEN_FILE.read_text(encoding="utf-8"))['expires_at'] - time.time()) / 3600)}h)")
        else:
            token, node_id = self._login(username, password)

        self.session.headers["Authorization"] = f"Bearer {token}"
        # Spec is internally inconsistent: descriptions say X-Node-Id, parameter says X-Node.
        # node_id=0 means no node routing needed, but send both defensively when present.
        if node_id:
            self.session.headers["X-Node"]    = str(node_id)
            self.session.headers["X-Node-Id"] = str(node_id)

    def _login(self, username, password):
        url = f"{self.host}/api/v3/auth/token"
        print(f"  Auth: POST {url}")
        r = requests.post(url, json={"username": username, "password": password}, timeout=TIMEOUT)
        body = r.json() if r.status_code == 200 else {}
        token = body.get("token")
        if not token:
            print(f"  [FATAL] auth failed (HTTP {r.status_code}): {r.text[:200]}")
            sys.exit(1)
        node_id = body.get("node_id", 0)
        save_token(token, node_id, body.get("expires_in", 172800))
        print(f"  Auth: login successful ✓  (node_id={node_id})")
        return token, node_id

    def get(self, path, params=None):
        url = f"{self.host}{path}"
        r = self.session.get(url, params=params, timeout=TIMEOUT)
        if r.status_code == 401:
            # Token may have been revoked; clear cache and re-login once
            print("  [warn] 401 — token may be stale, retrying with fresh login")
            try: TOKEN_FILE.unlink()
            except OSError: pass
            token, node_id = self._login(PILOT_USER, PILOT_PASS)
            self.session.headers["Authorization"] = f"Bearer {token}"
            if node_id:
                self.session.headers["X-Node"]    = str(node_id)
                self.session.headers["X-Node-Id"] = str(node_id)
            r = self.session.get(url, params=params, timeout=TIMEOUT)
        try:
            return r.status_code, r.json()
        except ValueError:
            return r.status_code, {"raw": r.text[:400]}

    def get_vehicles(self):
        status, body = self.get("/api/v3/vehicles")
        if status != 200 or body.get("code") != 0:
            print(f"  [FATAL] /vehicles failed: {body}")
            sys.exit(1)
        return body.get("data", []) or []

    def get_status_batch(self, imeis):
        if not imeis:
            return {}
        csv = ",".join(imeis)
        status, body = self.get("/api/v3/vehicles/status", params={"imei": csv})
        if status != 200 or body.get("code") != 0:
            print(f"  [warn] /vehicles/status failed: {body}")
            return {}
        return {str(v.get("imei")): v for v in body.get("data", []) or []}

    def get_distance(self, agent_id, ts, te):
        status, body = self.get("/api/v3/vehicles/distance",
                                params={"agent_id": agent_id, "ts": ts, "te": te})
        if status != 200 or body.get("code") != 0:
            return None
        return body.get("distance", {}) or {}


# ════════════════════════════════════════════════════════════════════════
# DATA NORMALIZATION
# ════════════════════════════════════════════════════════════════════════
def normalize_plate(raw):
    """Trim, smart-case, collapse spaces, replace underscores, split letter-digit runs."""
    if not raw:
        return ""
    s = raw.strip()
    # Underscores → spaces ("CADILLAC_2025_Black" → "CADILLAC 2025 Black")
    s = s.replace("_", " ")
    # Collapse multiple internal spaces
    s = re.sub(r"\s+", " ", s)
    # Insert space between digit and a run of 3+ lowercase letters
    # ("B15789mustang" → "B15789 mustang"; preserves "BMW420i" / "Q3" / "K5")
    s = re.sub(r"(\d)([a-z]{3,})", r"\1 \2", s)
    # Remove " -" / "- " artifacts mid-name ("L 94545 -Range" → "L 94545 Range")
    s = re.sub(r"\s-(?=\S)", " ", s)
    s = re.sub(r"(?<=\S)-\s", " ", s)
    # Per-word smart case for all-upper or all-lower alpha words ≥4 chars
    # (skips short codes like "BMW", "G63", "RS" — stays uppercase)
    parts = []
    for w in s.split(" "):
        if len(w) >= 4 and w.isalpha() and (w.isupper() or w.islower()):
            parts.append(w.title())
        else:
            parts.append(w)
    return " ".join(parts)

def is_sold(name):
    if not name:
        return False
    return name.strip().lower().startswith(SOLD_TOKENS)

def classify_zone(zone_field):
    """
    Return ('zone_text', 'category') where category is 'dubai', 'outside_dubai', or 'unknown'.
    The v3 /vehicles/status `zone` field is a dict {zone_id: zone_name}, sometimes empty list/None.
    Some deployments may return a list of strings — handle both.
    """
    if not zone_field:
        return "", "unknown"
    # Extract zone names from whichever shape we got
    if isinstance(zone_field, dict):
        names = [str(v) for v in zone_field.values() if v]
    elif isinstance(zone_field, list):
        names = [(str(v) if not isinstance(v, dict) else next(iter(v.values()), ""))
                 for v in zone_field if v]
    else:
        names = [str(zone_field)]
    if not names:
        return "", "unknown"
    # Title-case for readability (ALQOUZ INDUSTRIAL AREA 3 → Alqouz Industrial Area 3)
    pretty = [n.title() if n.isupper() else n for n in names]
    text = ", ".join(pretty)
    lower = text.lower()
    for kw in OUTSIDE_DUBAI_KEYWORDS:
        if kw in lower:
            return text, "outside_dubai"
    return text, "dubai"

def detect_engine_block(sensors):
    """Scan all sensors for any indication of engine block / immobilizer engaged."""
    if not sensors:
        return False
    for s in sensors:
        name = str(s.get("name", "")).lower()
        if "block" in name or "immobil" in name or "kill" in name:
            # Sensor exists — check value
            hum = str(s.get("hum_value", "")).lower()
            dig = s.get("dig_value")
            if hum in ("on", "blocked", "engaged", "1") or dig in (1, 1.0, True):
                return True
    return False

def find_sensor_value(sensors, *name_substrings):
    """Return first sensor's hum_value whose name contains any of the substrings."""
    if not sensors:
        return ""
    for s in sensors:
        name = str(s.get("name", "")).lower()
        for sub in name_substrings:
            if sub in name:
                return str(s.get("hum_value", ""))
    return ""

def fmt_hours_or_days(hours):
    """Format a duration in hours as either 'X.X hrs' or 'X.X days'."""
    if hours is None:
        return "—"
    if hours < 48:
        return f"{hours:.1f} hrs"
    return f"{hours / 24:.1f} days"

def fmt_ts(ts):
    try:
        return datetime.fromtimestamp(int(ts), tz=DUBAI).strftime("%Y-%m-%d %H:%M")
    except (TypeError, ValueError):
        return ""

def fmt_relative(ts, now=None):
    """'5m ago' / '2h ago' / '3d ago'."""
    if not ts:
        return "—"
    now = now or time.time()
    delta = max(0, int(now - int(ts)))
    if delta < 60:
        return "now"
    if delta < 3600:
        return f"{delta // 60}m ago"
    if delta < 86400:
        return f"{delta // 3600}h ago"
    return f"{delta // 86400}d ago"


# ════════════════════════════════════════════════════════════════════════
# RISK CLASSIFICATION
# ════════════════════════════════════════════════════════════════════════
def classify_risk(v):
    """
    Return one of 'Urgent', 'Watch', 'Normal'.
    Highest tier wins; reasons stored on the vehicle for optional Excel display.
    """
    reasons = []
    tier = "Normal"

    age_h = v.get("age_hours", 0)
    daily_km = v.get("daily_km", 0) or 0
    max_speed = v.get("max_speed", 0) or 0
    trips_cnt = v.get("trips_cnt", 0) or 0
    parking_h = v.get("parking_hrs", 0) or 0
    zone_cat = v.get("zone_category", "unknown")

    # ── URGENT ────────────────────────────────────────────────────────
    if age_h > URGENT_OFFLINE_HOURS:
        reasons.append(f"Offline {age_h:.0f}h")
        tier = "Urgent"
    if max_speed > URGENT_SPEED_KMH:
        reasons.append(f"Peak {max_speed} km/h")
        tier = "Urgent"
    if daily_km > URGENT_DAILY_KM:
        reasons.append(f"{daily_km:.0f} km driven")
        tier = "Urgent"
    if zone_cat == "outside_dubai":
        reasons.append("Outside Dubai")
        tier = "Urgent"
    if parking_h > 24 and zone_cat == "unknown" and trips_cnt == 0:
        reasons.append("Parked >24h, no zone")
        tier = "Urgent"

    # ── WATCH (only if not already Urgent) ────────────────────────────
    if tier != "Urgent":
        if WATCH_OFFLINE_MINUTES / 60.0 < age_h <= URGENT_OFFLINE_HOURS:
            reasons.append(f"Last update {age_h:.1f}h ago")
            tier = "Watch"
        if 130 < max_speed <= URGENT_SPEED_KMH:
            reasons.append(f"Peak {max_speed} km/h")
            tier = "Watch"
        if zone_cat == "unknown" and v.get("online"):
            reasons.append("Unknown zone")
            tier = "Watch"
        if trips_cnt > 0 and daily_km == 0:
            reasons.append("Moved but 0 km")
            tier = "Watch"
        if WATCH_DAILY_KM_MIN < daily_km <= URGENT_DAILY_KM:
            reasons.append(f"{daily_km:.0f} km driven")
            tier = "Watch"
        if parking_h > WATCH_PARKING_HOURS and zone_cat == "unknown":
            reasons.append(f"Parked {parking_h:.0f}h, no zone")
            tier = "Watch"

    return tier, reasons


# ════════════════════════════════════════════════════════════════════════
# VEHICLE ASSEMBLY
# ════════════════════════════════════════════════════════════════════════
def build_vehicle_records(client, fleet, report_date):
    """
    Merge /vehicles (inventory), /vehicles/status (live state, batch), and
    /vehicles/distance (per-vehicle, yesterday's window) into normalized records.
    """
    # ── Time window for "yesterday" in Dubai time ─────────────────────
    today_dubai = report_date.astimezone(DUBAI).replace(
        hour=0, minute=0, second=0, microsecond=0)
    ts = int((today_dubai - timedelta(days=1)).timestamp())
    te = int(today_dubai.timestamp())
    today_start_ts = int(today_dubai.timestamp())

    # ── Filter SOLD ───────────────────────────────────────────────────
    active = [v for v in fleet if not is_sold(v.get("vehiclenumber"))]
    sold_count = len(fleet) - len(active)
    print(f"  Fleet: {len(fleet)} total, {len(active)} active, {sold_count} excluded (SOLD)")

    # ── Batch status for live state ───────────────────────────────────
    imeis = [str(v["imei"]) for v in active if v.get("imei")]
    print(f"  Fetching /vehicles/status for {len(imeis)} IMEIs (batch)...")
    status_by_imei = client.get_status_batch(imeis)
    print(f"  Got status for {len(status_by_imei)} vehicles")

    # ── Per-vehicle /distance for yesterday ───────────────────────────
    print(f"  Fetching /vehicles/distance for {len(active)} vehicles (per-vehicle)...")
    snap = load_snapshot()
    last_move_history = snap.get("last_move_ts", {})

    records = []
    now = time.time()
    for i, v in enumerate(active, 1):
        agent_id = v.get("agentid")
        imei = str(v.get("imei", ""))
        st = status_by_imei.get(imei, {})
        dist = client.get_distance(agent_id, ts, te) or {}

        # ── Derived fields ────────────────────────────────────────────
        plate_raw = v.get("vehiclenumber", "")
        plate = normalize_plate(plate_raw)

        unix_ts = int(st.get("unixtimestamp", 0) or 0)
        age_seconds = max(0, now - unix_ts) if unix_ts else 99 * 3600
        age_hours = age_seconds / 3600.0
        online = age_hours < 0.5  # within last 30 minutes

        zone_list = st.get("zone", []) or []
        zone_text, zone_cat = classify_zone(zone_list)

        sensors = st.get("sensors", []) or []
        firing = int(st.get("firing", 0) or 0)
        ignition_on = firing == 1
        engine_blocked = detect_engine_block(sensors)
        car_batt = find_sensor_value(sensors, "car batt", "battery sensor")

        speed_now = float(st.get("speed", 0) or 0)
        is_moving = online and speed_now > 0

        # /distance fields (yesterday's window) — `length` is in km (see DISTANCE_UNIT_DIVISOR)
        length_km = round(int(dist.get("length", 0) or 0) / DISTANCE_UNIT_DIVISOR, 1)
        daily_km = length_km
        max_speed = int(dist.get("max_speed", 0) or 0)
        avg_speed = int(dist.get("avg_speed", 0) or 0)
        trips_cnt = int(dist.get("trips_cnt", 0) or 0)
        trips_te = int(dist.get("trips_te", 0) or 0)
        parking_time_s = int(dist.get("parking_time", 0) or 0)

        # Last-move timestamp: prefer yesterday's trips_te, else snapshot history
        last_move_ts = trips_te if trips_te > 0 else last_move_history.get(str(agent_id))
        if last_move_ts:
            parking_hrs = max(0.0, (now - int(last_move_ts)) / 3600.0)
        else:
            # No known last move — vehicle has been parked at least since snapshot start.
            # Use a conservative placeholder so first runs aren't misleading.
            parking_hrs = None

        # Tighter definition: a vehicle "moved today" only if it actually covered distance.
        # Pilot registers brief idle/sensor events as trips_cnt>0 with length=0 — those
        # should not count as movement. Without this, MSG2 fills with ⚠ 0 km noise.
        moved_today = trips_cnt > 0 and length_km > 0 and trips_te >= today_start_ts - 86400

        # Total odometer — prefer fresh current_mileage from /vehicles list
        total_km = round(float(v.get("current_mileage", 0) or 0), 1)

        rec = {
            "agent_id":       agent_id,
            "imei":           imei,
            "plate_raw":      plate_raw,
            "plate":          plate,
            "folder":         v.get("folder", ""),
            "total_km":       total_km,
            "online":         online,
            "age_hours":      age_hours,
            "is_moving":      is_moving,
            "moved_today":    moved_today,
            "speed":          speed_now,
            "ignition_on":    ignition_on,
            "engine_blocked": engine_blocked,
            "car_batt":       car_batt,
            "lat":            st.get("lat"),
            "lon":            st.get("lon"),
            "unixtimestamp":  unix_ts,
            "last_seen":      fmt_ts(unix_ts) if unix_ts else "",
            "zone":           zone_text,
            "zone_category":  zone_cat,
            "daily_km":       daily_km,
            "max_speed":      max_speed,
            "avg_speed":      avg_speed,
            "trips_cnt":      trips_cnt,
            "trips_te":       trips_te,
            "last_move_ts":   last_move_ts,
            "last_move":      fmt_ts(last_move_ts) if last_move_ts else "",
            "parking_secs":   parking_time_s,
            "parking_hrs":    parking_hrs if parking_hrs is not None else 0,
            "parking_hrs_known": parking_hrs is not None,
        }
        risk, reasons = classify_risk(rec)
        rec["risk"] = risk
        rec["risk_reasons"] = reasons
        records.append(rec)

        if i % 10 == 0:
            print(f"    ...{i}/{len(active)}")

    # ── First-run / long-idle fallback ──────────────────────────────
    # For vehicles with no known last-move (didn't move yesterday AND no snapshot history),
    # query /distance with a 30-day window to recover trips_te. This avoids the "all
    # parked vehicles show — for idle time" first-run problem.
    needs_fallback = [r for r in records if not r["parking_hrs_known"]]
    if needs_fallback:
        print(f"  Recovering last-move for {len(needs_fallback)} parked vehicles (30-day window)...")
        ts_30d = int(now - 30 * 86400)
        te_now = int(now)
        for r in needs_fallback:
            ext = client.get_distance(r["agent_id"], ts_30d, te_now) or {}
            te_recent = int(ext.get("trips_te", 0) or 0)
            if te_recent > 0:
                r["last_move_ts"] = te_recent
                r["last_move"] = fmt_ts(te_recent)
                r["parking_hrs"] = max(0.0, (now - te_recent) / 3600.0)
                r["parking_hrs_known"] = True
                # Re-classify risk in case parking_hrs now triggers a rule
                risk, reasons = classify_risk(r)
                r["risk"] = risk
                r["risk_reasons"] = reasons

    # Sanity check for unit drift
    moved_recs = [r for r in records if r["moved_today"]]
    total_km = sum(r["daily_km"] for r in moved_recs)
    if len(moved_recs) > 10 and total_km < 100:
        print(f"  [WARN] total daily km is {total_km:.1f} across {len(moved_recs)} moved vehicles. "
              f"Suspect unit drift — check DISTANCE_UNIT_DIVISOR (currently {DISTANCE_UNIT_DIVISOR}).")

    return records, sold_count


# ════════════════════════════════════════════════════════════════════════
# SLACK MESSAGES
# ════════════════════════════════════════════════════════════════════════
def _truncate(s, n):
    """Truncate to n chars, preferring a word boundary if it's within ~8 chars of the cut."""
    s = s or ""
    if len(s) <= n:
        return s
    cut = s[:n - 1]
    last_space = cut.rfind(" ")
    if last_space >= n - 9:
        cut = cut[:last_space]
    return cut + "…"

def build_msg1_snapshot(vehicles, sold_count, date_str):
    online   = sum(1 for v in vehicles if v["online"])
    offline  = sum(1 for v in vehicles if not v["online"])
    moved    = sum(1 for v in vehicles if v["moved_today"])
    total_km = sum(v["daily_km"] for v in vehicles)
    in_zone      = sum(1 for v in vehicles if v["zone_category"] == "dubai")
    outside_dxb  = sum(1 for v in vehicles if v["zone_category"] == "outside_dubai")
    blocked  = sum(1 for v in vehicles if v["engine_blocked"])
    urgent_n = sum(1 for v in vehicles if v["risk"] == "Urgent")
    watch_n  = sum(1 for v in vehicles if v["risk"] == "Watch")
    attn_n   = urgent_n + watch_n
    total    = len(vehicles)

    msg = (
        f":car: *MKV CAR RENTAL — Daily GPS Report*\n"
        f"*Date:* {date_str}  |  *Timezone:* {TIMEZONE_NAME}\n"
        f"─────────────────────────────\n"
        f":white_check_mark: *Online:* {online} / {total}\n"
        f":red_circle: *Offline:* {offline}\n"
        f":chart_with_upwards_trend: *Moved Today:* {moved} vehicles\n"
        f":round_pushpin: *Total Km Driven Today:* {total_km:,.1f} km\n"
        f":rotating_light: *Needs Attention:* {attn_n}  "
        f"({urgent_n} Urgent · {watch_n} Watch)\n"
        f":busts_in_silhouette: *In a Zone:* {in_zone}  ·  "
        f":triangular_flag_on_post: *Outside Dubai:* {outside_dxb}\n"
        f":lock: *Engine Blocked:* {blocked}\n"
        f"─────────────────────────────\n"
    )
    if sold_count:
        msg += f"_Excludes {sold_count} SOLD vehicle{'s' if sold_count != 1 else ''}._  "
    msg += f"_Generated at {datetime.now(DUBAI).strftime('%Y-%m-%d %H:%M')} (Dubai time)._"
    return msg


def build_msg2_moved(vehicles):
    moved = [v for v in vehicles if v["moved_today"]]
    moved.sort(key=lambda x: x["plate"].lower())

    lines = [f"{'Vehicle':<38} {'Km Today':>10}   Zone"]
    lines.append("─" * 81)
    for v in moved:
        km_str = f"{v['daily_km']:.1f} km"
        # 🚩 marker for outside-Dubai endings
        flag = "🚩 " if v["zone_category"] == "outside_dubai" else ""
        zone_disp = (flag + v["zone"]) if v["zone"] else "—"
        lines.append(f"{_truncate(v['plate'], 37):<38} {km_str:>10}   {_truncate(zone_disp, 35)}")
    body = "\n".join(lines)

    return (
        f":blue_car: *Vehicles That Moved Today ({len(moved)})*\n"
        f"```\n{body}\n```"
    )


def build_msg3_parked(vehicles):
    # Only online + didn't move today
    parked = [v for v in vehicles if not v["moved_today"] and v["online"]]
    offline_count = sum(1 for v in vehicles if not v["online"])

    short_idle = [v for v in parked if v["parking_hrs_known"] and v["parking_hrs"] < LONG_IDLE_HOURS]
    long_idle  = [v for v in parked if not v["parking_hrs_known"] or v["parking_hrs"] >= LONG_IDLE_HOURS]
    short_idle.sort(key=lambda x: x["parking_hrs"], reverse=True)
    long_idle.sort(key=lambda x: (x["parking_hrs"] if x["parking_hrs_known"] else 0), reverse=True)

    def table(rows, header_units):
        lines = [f"{'Vehicle':<38} {header_units:>11}   Zone"]
        lines.append("─" * 81)
        for v in rows:
            if not v["parking_hrs_known"]:
                dur = "—"
            else:
                dur = fmt_hours_or_days(v["parking_hrs"])
            zone_disp = v["zone"] if v["zone"] else "—"
            lines.append(f"{_truncate(v['plate'], 37):<38} {dur:>11}   {_truncate(zone_disp, 35)}")
        return "\n".join(lines)

    parts = [
        f":parking: *Idle <48h ({len(short_idle)} vehicles)*\n"
        f"```\n{table(short_idle, 'Idle Time')}\n```"
    ]
    parts.append(
        f"\n:package: *Long-idle >48h ({len(long_idle)} vehicles)* — review for utilization\n"
        f"```\n{table(long_idle, 'Idle Time')}\n```"
    )
    if offline_count:
        parts.append(f"\n_:red_circle: {offline_count} vehicle{'s' if offline_count != 1 else ''} "
                     f"offline — excluded from list._")
    return "".join(parts)


def build_msg4_speeding(vehicles):
    # Aggregate per vehicle: any vehicle with max_speed > threshold gets a row.
    # "Events" count: we don't have per-event from v3 API, so we report 1 if max_speed > limit
    # (representing "at least one event with peak X km/h yesterday"). Most fleets care about
    # the peak and frequency proxy more than exact event count.
    offenders = [v for v in vehicles if (v["max_speed"] or 0) > SPEED_LIMIT_KMH]
    offenders.sort(key=lambda x: (x["max_speed"], x["avg_speed"]), reverse=True)

    if not offenders:
        return f":white_check_mark: *No speeding events recorded yesterday (limit {SPEED_LIMIT_KMH} km/h)*"

    lines = [f"{'Vehicle':<38} {'Peak':>5}  {'Avg':>5}  Trips"]
    lines.append("─" * 68)
    for v in offenders:
        lines.append(
            f"{_truncate(v['plate'], 37):<38} "
            f"{v['max_speed']:>5}  {v['avg_speed']:>5}  {v['trips_cnt']:>5}"
        )
    body = "\n".join(lines)
    return (
        f":rotating_light: *Speeding Yesterday — {len(offenders)} vehicle"
        f"{'s' if len(offenders) != 1 else ''} over {SPEED_LIMIT_KMH} km/h*\n"
        f"```\n{body}\n```"
    )


def post_slack(messages):
    if not SLACK_TOKEN:
        print("  [Slack] No token — skipping post")
        for i, m in enumerate(messages, 1):
            print(f"\n──── MSG{i} ────")
            print(m)
        return
    headers = {"Authorization": f"Bearer {SLACK_TOKEN}", "Content-Type": "application/json"}
    ok = 0
    for m in messages:
        try:
            r = requests.post("https://slack.com/api/chat.postMessage",
                              headers=headers,
                              json={"channel": SLACK_CHANNEL, "text": m},
                              timeout=10)
            result = r.json()
            if result.get("ok"):
                ok += 1
            else:
                print(f"  [Slack] Failed: {result.get('error')}")
        except requests.RequestException as e:
            print(f"  [Slack] Error: {e}")
    print(f"  [Slack] Posted {ok}/{len(messages)} messages to #{SLACK_CHANNEL} ✓")


# ════════════════════════════════════════════════════════════════════════
# EXCEL — preserves 5-tab structure from original
# ════════════════════════════════════════════════════════════════════════
def build_summary(wb, date_str, vehicles, sold_count):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25

    ws.merge_cells("B1:C1")
    t = ws["B1"]; t.value = "MKV CAR RENTAL — Daily GPS Report"
    t.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    t.fill = hfill(DARK_GREEN); t.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("B2:C2")
    s = ws["B2"]; s.value = (f"{date_str}  |  {TIMEZONE_NAME}  |  "
                             f"{datetime.now(DUBAI).strftime('%Y-%m-%d %H:%M')}")
    s.font = Font(name="Arial", size=10, color=WHITE)
    s.fill = hfill(MID_GREEN); s.alignment = center()
    ws.row_dimensions[2].height = 20

    online   = sum(1 for v in vehicles if v["online"])
    moved    = sum(1 for v in vehicles if v["moved_today"])
    blocked  = sum(1 for v in vehicles if v["engine_blocked"])
    in_zone  = sum(1 for v in vehicles if v["zone_category"] == "dubai")
    outside  = sum(1 for v in vehicles if v["zone_category"] == "outside_dubai")
    urgent   = sum(1 for v in vehicles if v["risk"] == "Urgent")
    watch    = sum(1 for v in vehicles if v["risk"] == "Watch")
    total_km     = sum(v["total_km"] for v in vehicles)
    total_daily  = sum(v["daily_km"] for v in vehicles)

    ws["B4"] = "Fleet Summary"
    ws["B4"].font = Font(name="Arial", bold=True, size=12, color=DARK_GREEN)

    stats = [
        ("Total Active Vehicles",  len(vehicles)),
        ("Excluded (SOLD)",        sold_count),
        ("Online Now",             f"{online} ✓"),
        ("Offline",                len(vehicles) - online),
        ("Moved Today",            f"{moved} vehicles"),
        ("Urgent",                 urgent),
        ("Watch",                  watch),
        ("In Dubai Zone",          in_zone),
        ("Outside Dubai",          outside),
        ("Engine Blocked",         blocked),
        ("Km Driven Yesterday",    f"{total_daily:,.1f} km"),
        ("Total Fleet Odometer",   f"{total_km:,.1f} km"),
    ]
    for row_i, (label, val) in enumerate(stats, 5):
        ws.cell(row_i, 2).value = label
        ws.cell(row_i, 2).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row_i, 2).fill = PatternFill("solid", fgColor=LIGHT_GREY)
        ws.cell(row_i, 3).value = str(val)
        ws.cell(row_i, 3).font = Font(name="Arial", size=10)
        ws.row_dimensions[row_i].height = 18

    start_tabs = 5 + len(stats) + 1
    ws.cell(start_tabs, 2).value = "Report Tabs"
    ws.cell(start_tabs, 2).font = Font(name="Arial", bold=True, size=12, color=DARK_GREEN)
    tabs = [
        ("Daily Movement",  "Km yesterday, last move/stop, moved today"),
        ("Activity Report", "Status, ignition, parking duration, zone"),
        ("Speed & Status",  "Speed, battery, engine block, online/offline"),
        ("Geofence",        "Current zone per vehicle, outside-Dubai flag"),
    ]
    for i, (tab, desc) in enumerate(tabs, 1):
        ws.cell(start_tabs + i, 2).value = tab
        ws.cell(start_tabs + i, 2).font = Font(name="Arial", bold=True, size=10, color=MID_GREEN)
        ws.cell(start_tabs + i, 3).value = desc
        ws.row_dimensions[start_tabs + i].height = 18


def build_daily_movement(wb, vehicles, date_str):
    ws = wb.create_sheet("Daily Movement")
    headers = ["#", "Vehicle Name", "Moved Today", "Km Yesterday", "Max Speed",
               "Trips", "Last Move", "Current Status", "Online", "Zone", "Risk"]
    widths  = [5, 32, 13, 14, 11, 7, 20, 14, 10, 30, 10]
    title_block(ws, "DAILY MOVEMENT REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        status   = "Moving" if v["is_moving"] else "Parked" if v["online"] else "Offline"
        ws.append([i, v["plate"],
                   "✓ Yes" if v["moved_today"] else "No",
                   v["daily_km"], v["max_speed"], v["trips_cnt"],
                   v["last_move"], status,
                   "Online" if v["online"] else "Offline",
                   v["zone"] or "—", v["risk"]])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))
        # Risk color
        rc = ws.cell(data_start + i - 1, 11)
        if v["risk"] == "Urgent":
            rc.font = Font(name="Arial", bold=True, size=10, color=RED)
        elif v["risk"] == "Watch":
            rc.font = Font(name="Arial", bold=True, size=10, color=AMBER)
        else:
            rc.font = Font(name="Arial", size=10, color=GREY)
    ws.freeze_panes = "A5"


def build_activity(wb, vehicles, date_str):
    ws = wb.create_sheet("Activity Report")
    headers = ["#", "Vehicle Name", "Status", "Ignition",
               "Parking Duration", "Last Seen", "Zone", "Online"]
    widths  = [5, 32, 12, 10, 18, 20, 30, 10]
    title_block(ws, "ACTIVITY REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        status = "Moving" if v["is_moving"] else "Parked" if v["online"] else "Offline"
        p_disp = (fmt_hours_or_days(v["parking_hrs"]) if v["parking_hrs_known"] and not v["is_moving"]
                  else "—")
        ws.append([i, v["plate"], status,
                   "ON" if v["ignition_on"] else "off",
                   p_disp, v["last_seen"], v["zone"] or "—",
                   "Online" if v["online"] else "Offline"])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))
    ws.freeze_panes = "A5"


def build_speed_status(wb, vehicles, date_str):
    ws = wb.create_sheet("Speed & Status")
    headers = ["#", "Vehicle Name", "Online", "Current Speed",
               "Max Speed Yest.", "Engine Block", "Battery", "Ignition", "Last Seen"]
    widths  = [5, 32, 10, 14, 16, 14, 16, 10, 20]
    title_block(ws, "SPEED & STATUS REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        ws.append([i, v["plate"],
                   "Online" if v["online"] else "Offline",
                   v["speed"], v["max_speed"],
                   "BLOCKED" if v["engine_blocked"] else "OK",
                   v["car_batt"],
                   "ON" if v["ignition_on"] else "off",
                   v["last_seen"]])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))
        if v["max_speed"] > URGENT_SPEED_KMH:
            ws.cell(data_start+i-1, 5).font = Font(name="Arial", bold=True, color=RED)
        elif v["max_speed"] > SPEED_LIMIT_KMH:
            ws.cell(data_start+i-1, 5).font = Font(name="Arial", bold=True, color=AMBER)
    ws.freeze_panes = "A5"


def build_geofence(wb, vehicles, date_str):
    ws = wb.create_sheet("Geofence")
    headers = ["#", "Vehicle Name", "Current Zone", "Zone Category",
               "Status", "Online", "Last Seen"]
    widths  = [5, 32, 40, 14, 12, 10, 20]
    title_block(ws, "GEOFENCE REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        cat_label = {"dubai": "Dubai", "outside_dubai": "🚩 Outside Dubai",
                     "unknown": "Unknown"}[v["zone_category"]]
        ws.append([i, v["plate"], v["zone"] or "Outside all zones", cat_label,
                   "Moving" if v["is_moving"] else "Parked",
                   "Online" if v["online"] else "Offline", v["last_seen"]])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))
        if v["zone_category"] == "outside_dubai":
            ws.cell(data_start+i-1, 4).font = Font(name="Arial", bold=True, color=RED)
    ws.freeze_panes = "A5"


# ════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="Report date YYYY-MM-DD (default: today)")
    parser.add_argument("--dry-run", action="store_true", help="Skip Slack post, print messages")
    args = parser.parse_args()

    if not PILOT_USER or not PILOT_PASS:
        print("ERROR: PILOT_USER and PILOT_PASS must be set in .env")
        sys.exit(1)

    report_date = datetime.now(DUBAI)
    if args.date:
        report_date = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=DUBAI)
    date_str = report_date.strftime("%Y-%m-%d")

    print(f"\n{'='*55}")
    print(f"  MKV CAR RENTAL — Daily GPS Report (v3 API)")
    print(f"  Date: {date_str}   Host: {PILOT_HOST}")
    print(f"{'='*55}\n")

    client = PilotV3Client(PILOT_HOST, PILOT_USER, PILOT_PASS)

    print("\nFetching fleet inventory...")
    fleet = client.get_vehicles()

    print("Assembling vehicle records...")
    vehicles, sold_count = build_vehicle_records(client, fleet, report_date)

    online   = sum(1 for v in vehicles if v["online"])
    moved    = sum(1 for v in vehicles if v["moved_today"])
    urgent   = sum(1 for v in vehicles if v["risk"] == "Urgent")
    watch    = sum(1 for v in vehicles if v["risk"] == "Watch")
    total_km = sum(v["daily_km"] for v in vehicles)
    print(f"\n  Active: {len(vehicles)}  Online: {online}  Moved today: {moved}")
    print(f"  Total km yesterday: {total_km:,.1f}")
    print(f"  Risk: {urgent} Urgent · {watch} Watch · {len(vehicles)-urgent-watch} Normal")

    # ── Excel ─────────────────────────────────────────────────────────
    print("\nBuilding Excel workbook...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    build_summary(wb, date_str, vehicles, sold_count)
    build_daily_movement(wb, vehicles, date_str)
    build_activity(wb, vehicles, date_str)
    build_speed_status(wb, vehicles, date_str)
    build_geofence(wb, vehicles, date_str)
    fname = f"MKV_GPS_Report_{date_str}.xlsx"
    out = SCRIPT_DIR / fname
    wb.save(out)
    print(f"  Saved: {fname}")

    # ── Snapshot for next run ─────────────────────────────────────────
    save_snapshot(vehicles, date_str)

    # ── Slack ─────────────────────────────────────────────────────────
    print("\nBuilding Slack messages...")
    messages = [
        build_msg1_snapshot(vehicles, sold_count, date_str),
        build_msg2_moved(vehicles),
        build_msg3_parked(vehicles),
        build_msg4_speeding(vehicles),
    ]
    if args.dry_run:
        print("  [dry-run] not posting; messages would be:")
        for i, m in enumerate(messages, 1):
            print(f"\n──── MSG{i} ────")
            print(m)
    else:
        post_slack(messages)

    print(f"\n{'='*55}\n")


if __name__ == "__main__":
    main()
