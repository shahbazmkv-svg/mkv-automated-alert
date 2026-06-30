"""
MKV CAR RENTAL - GPS2: AL ETIQAN Daily Report
PRODUCTION VERSION - Built from real API test data (2026-06-29)

Key design decisions:
- Snapshot key: Registration plate code (e.g., "U-74545", "V-1243", "D-68539")
- Handles asterisks in vehicle names (** separator pattern)
- Status mapping: moving/stopped/idle properly recognized
- 4 Slack messages + Excel workbook matching GPS3 format
"""

import requests
import json
import os
import sys
import re
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv

load_dotenv()

# ─── CONFIG ─────────────────────────────────────────
ETIQAN_BASE   = "http://track.etqanuae.com/api/api.php"
ETIQAN_KEY    = os.getenv("ETIQAN_API_KEY", "8651C0D3A56F60178A09B3007B8BF32B")
SLACK_TOKEN   = os.getenv("SLACK_TOKEN", "")
SLACK_CHANNEL = os.getenv("SLACK_CHANNEL", "C0B6Y6EG85D")
SNAPSHOT_FILE = "gps2_etiqan_snapshot.json"
DUBAI_OFFSET  = timezone(timedelta(hours=4))
TIMEZONE_NAME = "Asia/Dubai"

# Style colors (matching GPS3 format)
COLOR_NAVY   = "1B2A5A"
COLOR_GOLD   = "C9A84C"
COLOR_RED    = "C0392B"
COLOR_GREEN  = "27AE60"


# ─── HELPERS ────────────────────────────────────────
def parse_vehicle_name(name):
    """
    Extract device_id (registration plate) and clean display name.
    Handles ETIQAN's asterisk naming convention.
    
    Examples:
      "U-74545 Ferrari"           → ("U-74545", "U-74545 Ferrari")
      "V-1243**Rolls Royce Culli" → ("V-1243", "V-1243 Rolls Royce Culli")
      "D-68539**Forthing S7"      → ("D-68539", "D-68539 Forthing S7")
    """
    # device_id: split on ** OR whitespace, take first part
    device_id = re.split(r'\*\*|\s', name, maxsplit=1)[0]
    
    # Display name: replace ** with single space, clean up
    display_name = re.sub(r'\*+', ' ', name).strip()
    display_name = re.sub(r'\s+', ' ', display_name)  # collapse multiple spaces
    
    return device_id, display_name


def normalize_status(status_raw):
    """Map ETIQAN status to standardized values."""
    status_raw = str(status_raw).lower().strip()
    if "moving" in status_raw:
        return "Moving"
    elif "idle" in status_raw:
        return "Idle"  # Engine on, not moving
    elif "stopped" in status_raw or "parked" in status_raw:
        return "Stopped"
    elif "offline" in status_raw or "no" in status_raw:
        return "Offline"
    else:
        return status_raw.title()


# ─── FETCH VEHICLES ─────────────────────────────────
def fetch_vehicles():
    """Get vehicles from ETIQAN API USER_GET_OBJECTS_STATUS endpoint."""
    params = {
        "api": "user", "ver": "1.0", "key": ETIQAN_KEY,
        "cmd": "USER_GET_OBJECTS_STATUS",
    }
    
    try:
        r = requests.get(ETIQAN_BASE, params=params, timeout=20)
        print(f"  [GPS2] API HTTP: {r.status_code}")
        
        if r.status_code != 200:
            print(f"  [GPS2] API error: {r.text[:200]}")
            return []
        
        # Check for error responses (e.g., "wrong API key" returned as plain text)
        if not r.text.strip().startswith('['):
            print(f"  [GPS2] API error response: {r.text[:200]}")
            return []
        
        data = r.json()
    except Exception as e:
        print(f"  [GPS2] API exception: {e}")
        return []
    
    if not isinstance(data, list):
        print(f"  [GPS2] Unexpected response format")
        return []
    
    vehicles = []
    for obj in data:
        raw_name = str(obj.get("name", "Unknown"))
        device_id, display_name = parse_vehicle_name(raw_name)
        status = normalize_status(obj.get("status", ""))
        
        try:
            speed_kmh = float(obj.get("speed", 0))
        except:
            speed_kmh = 0
        
        try:
            odometer_km = float(obj.get("odometer", 0))
        except:
            odometer_km = 0
        
        vehicles.append({
            "raw_name":    raw_name,        # Original API name (for overspeed lookup)
            "device_id":   device_id,       # Registration plate (snapshot key)
            "name":        display_name,    # Cleaned name (for display)
            "status":      status,
            "speed_kmh":   speed_kmh,
            "odometer_km": odometer_km,
            "daily_km":    0.0,
            "lat":         obj.get("lat", ""),
            "lng":         obj.get("lng", ""),
            "last_update": obj.get("dt_tracker", ""),
        })
        print(f"    {display_name:<35} [{status:<8}] speed={speed_kmh:>3} odo={odometer_km:>10,.1f} km")
    
    print(f"  [GPS2] Fetched {len(vehicles)} vehicles")
    return vehicles


# ─── OVERSPEED EVENTS ───────────────────────────────
def fetch_overspeed_events(vehicle_raw_name, date_from, date_to):
    """Fetch overspeed events for a specific vehicle."""
    cmd = f"OBJECT_GET_EVENTS_OVERSPEED,{vehicle_raw_name},{date_from} 00:00:00,{date_to} 23:59:59"
    params = {
        "api": "user", "ver": "1.0", "key": ETIQAN_KEY,
        "cmd": cmd,
    }
    
    try:
        r = requests.get(ETIQAN_BASE, params=params, timeout=20)
        if r.status_code != 200 or not r.text.strip().startswith('['):
            return []
        events = r.json()
        return events if isinstance(events, list) else []
    except:
        return []


# ─── SNAPSHOT MANAGEMENT ────────────────────────────
def load_snapshot():
    """Load yesterday's snapshot."""
    try:
        with open(SNAPSHOT_FILE, "r") as f:
            return json.load(f)
    except:
        return {}


def save_snapshot(vehicles, date_str):
    """Save snapshot using device_id (registration plate) as key."""
    snap = {
        "date": date_str,
        "odometers": {v["device_id"]: v["odometer_km"] for v in vehicles}
    }
    try:
        with open(SNAPSHOT_FILE, "w", encoding="utf-8") as f:
            json.dump(snap, f, indent=2, ensure_ascii=False)
        print(f"  [GPS2] Snapshot saved ({len(snap['odometers'])} vehicles)")
    except Exception as e:
        print(f"  [GPS2] Save error: {e}")


def calculate_deltas(vehicles, snapshot):
    """Calculate daily_km from odometer delta using device_id as key."""
    snap_date = snapshot.get("date", "")
    snap_odos = snapshot.get("odometers", {})
    
    today = datetime.now(DUBAI_OFFSET).strftime("%Y-%m-%d")
    yesterday = (datetime.now(DUBAI_OFFSET) - timedelta(days=1)).strftime("%Y-%m-%d")
    
    if snap_date not in (today, yesterday):
        print(f"  [GPS2] Snapshot date '{snap_date}' not recent — skipping deltas")
        return
    
    print(f"  [GPS2] Computing deltas from snapshot ({snap_date})")
    for v in vehicles:
        device_id = v["device_id"]
        curr = v["odometer_km"]
        prev = snap_odos.get(device_id)
        
        if prev is not None and curr > 0:
            v["daily_km"] = max(0.0, round(curr - prev, 1))
            print(f"    {v['name']:<35} {curr:>10,.1f} - {prev:>10,.1f} = {v['daily_km']:>6.1f} km")
        else:
            print(f"    {v['name']:<35} NO MATCH (device_id='{device_id}')")


# ─── SLACK MESSAGES ─────────────────────────────────
def build_slack_messages(vehicles, date_str, overspeed_data):
    """Build the 4 Slack messages."""
    total_km     = sum(v["daily_km"] for v in vehicles)
    moved        = [v for v in vehicles if v["daily_km"] > 0]
    parked       = [v for v in vehicles if v["daily_km"] == 0 and v["status"] != "Offline"]
    online       = sum(1 for v in vehicles if v["status"] != "Offline")
    offline      = len(vehicles) - online
    moving_now   = sum(1 for v in vehicles if v["status"] == "Moving")
    idle_now     = sum(1 for v in vehicles if v["status"] == "Idle")
    
    now = datetime.now(DUBAI_OFFSET).strftime("%Y-%m-%d %H:%M")
    
    # ── MSG 1: Summary ──
    msg1 = (
        f":etiqan: *MKV CAR RENTAL — GPS2: AL ETIQAN Daily Report*\n"
        f"*Date:* {date_str}  |  *Timezone:* {TIMEZONE_NAME}\n"
        f"─────────────────────────────\n"
        f":white_check_mark: *Online:* {online} / {len(vehicles)}\n"
        f":red_circle: *Offline:* {offline}\n"
        f":blue_car: *Moving Now:* {moving_now}\n"
        f":hourglass_flowing_sand: *Idle:* {idle_now}\n"
        f":chart_with_upwards_trend: *Moved Today:* {len(moved)} vehicles\n"
        f":round_pushpin: *Total Km Driven Today:* {total_km:.1f} km\n"
        f":busts_in_silhouette: *In a Zone:* —\n"
        f":lock: *Engine Blocked:* —\n"
        f"─────────────────────────────\n"
        f"_Generated at {now} (Dubai time)_"
    )
    
    # ── MSG 2: Vehicles Moved ──
    if moved:
        rows = "\n".join(f"{v['name']:<35} {v['daily_km']:>7.1f} km   —" for v in moved)
    else:
        rows = "No vehicles moved today"
    
    msg2 = (
        f":blue_car: *GPS2 — Vehicles That Moved Today ({len(moved)})*\n"
        f"```\n"
        f"Vehicle                              Km Today  Zone\n"
        f"───────────────────────────────────────────────────────────────\n"
        f"{rows}\n"
        f"```"
    )
    
    # ── MSG 3: Parked Vehicles ──
    if parked:
        rows = "\n".join(f"{v['name']:<35} {v['status']:>10}   —" for v in parked)
    else:
        rows = "No parked vehicles"
    
    msg3 = (
        f":parking: *GPS2 — Parked / Idle Vehicles ({len(parked)})*\n"
        f"```\n"
        f"Vehicle                              Status        Zone\n"
        f"───────────────────────────────────────────────────────────────\n"
        f"{rows}\n"
        f"```\n"
        f"_:red_circle: {offline} vehicle(s) offline_"
    )
    
    # ── MSG 4: Speeding Events ──
    total_events = sum(len(events) for events in overspeed_data.values())
    if total_events > 0:
        lines = []
        for v in vehicles:
            events = overspeed_data.get(v["raw_name"], [])
            if events:
                # Get max speed and event count
                max_speed = max(int(e.get("speed", 0)) for e in events)
                lines.append(f"{v['name']:<35} {len(events):>3} events  max {max_speed} km/h")
        rows = "\n".join(lines)
        msg4 = (
            f":rotating_light: *GPS2 — Speeding Events Yesterday ({total_events} total)*\n"
            f"```\n"
            f"Vehicle                              Events    Max Speed\n"
            f"───────────────────────────────────────────────────────────────\n"
            f"{rows}\n"
            f"```"
        )
    else:
        msg4 = ":white_check_mark: *GPS2 — No speeding events recorded yesterday*"
    
    return [msg1, msg2, msg3, msg4]


def post_to_slack(messages, dry_run=False):
    """Display messages and optionally post to Slack."""
    print("\n" + "─" * 60)
    print("  SLACK MESSAGE PREVIEW")
    print("─" * 60)
    for i, m in enumerate(messages, 1):
        print(f"\n── MSG {i} ──────────────────────")
        print(m)
    print("\n" + "─" * 60)
    
    if dry_run:
        print("  [DRY-RUN] Slack posting SKIPPED")
        return
    
    if not SLACK_TOKEN:
        print("  [GPS2] No SLACK_TOKEN, skipping Slack post")
        return
    
    headers = {"Authorization": f"Bearer {SLACK_TOKEN}", "Content-Type": "application/json"}
    posted = 0
    for i, m in enumerate(messages, 1):
        try:
            r = requests.post(
                "https://slack.com/api/chat.postMessage",
                headers=headers,
                json={"channel": SLACK_CHANNEL, "text": m},
                timeout=15,
            )
            if r.json().get("ok"):
                posted += 1
            else:
                print(f"  [Slack] MSG{i} error: {r.json().get('error')}")
        except Exception as e:
            print(f"  [Slack] MSG{i} exception: {e}")
    print(f"  [Slack] Posted {posted}/{len(messages)} messages")


# ─── EXCEL ──────────────────────────────────────────
def build_excel(vehicles, date_str, overspeed_data):
    """Generate Excel workbook with 4 tabs."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    nav_font = Font(bold=True, color="FFFFFF", name="Arial")
    nav_fill = PatternFill("solid", fgColor=COLOR_NAVY)
    gold_fill = PatternFill("solid", fgColor=COLOR_GOLD)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ── TAB 1: Summary ──
    ws = wb.create_sheet("GPS2 Summary", 0)
    ws.cell(1, 1, f"MKV CAR RENTAL — GPS2: AL ETIQAN").font = Font(bold=True, size=14)
    ws.cell(2, 1, f"Date: {date_str}").font = Font(italic=True)
    
    summary_rows = [
        ("Metric", "Value"),
        ("Total Vehicles", len(vehicles)),
        ("Online", sum(1 for v in vehicles if v["status"] != "Offline")),
        ("Offline", sum(1 for v in vehicles if v["status"] == "Offline")),
        ("Moving Now", sum(1 for v in vehicles if v["status"] == "Moving")),
        ("Idle", sum(1 for v in vehicles if v["status"] == "Idle")),
        ("Moved Today", sum(1 for v in vehicles if v["daily_km"] > 0)),
        ("Total Km Driven", f"{sum(v['daily_km'] for v in vehicles):.1f} km"),
    ]
    for r, (k, v) in enumerate(summary_rows, 4):
        c1 = ws.cell(r, 1, k)
        c2 = ws.cell(r, 2, v)
        if r == 4:
            c1.font = nav_font
            c2.font = nav_font
            c1.fill = nav_fill
            c2.fill = nav_fill
        c1.border = border
        c2.border = border
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    
    # ── TAB 2: Daily Movement ──
    ws = wb.create_sheet("GPS2 Daily Movement", 1)
    ws.cell(1, 1, "Daily Movement").font = Font(bold=True, size=14)
    headers = ["Vehicle", "Km Today", "Status", "Current Speed"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = nav_font
        cell.fill = nav_fill
        cell.border = border
    for r, v in enumerate(vehicles, 4):
        ws.cell(r, 1, v["name"])
        ws.cell(r, 2, v["daily_km"])
        ws.cell(r, 3, v["status"])
        ws.cell(r, 4, v["speed_kmh"])
        for c in range(1, 5):
            ws.cell(r, c).border = border
    for c in range(1, 5):
        ws.column_dimensions[chr(64 + c)].width = 22
    
    # ── TAB 3: Parked Vehicles ──
    ws = wb.create_sheet("GPS2 Parked Vehicles", 2)
    ws.cell(1, 1, "Parked / Idle Vehicles").font = Font(bold=True, size=14)
    headers = ["Vehicle", "Status", "Last Update", "Location"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = nav_font
        cell.fill = nav_fill
        cell.border = border
    parked = [v for v in vehicles if v["status"] != "Moving"]
    for r, v in enumerate(parked, 4):
        ws.cell(r, 1, v["name"])
        ws.cell(r, 2, v["status"])
        ws.cell(r, 3, v["last_update"])
        ws.cell(r, 4, f"{v['lat']}, {v['lng']}")
        for c in range(1, 5):
            ws.cell(r, c).border = border
    for c in range(1, 5):
        ws.column_dimensions[chr(64 + c)].width = 25
    
    # ── TAB 4: Fleet Status (with overspeed) ──
    ws = wb.create_sheet("GPS2 Fleet Status", 3)
    ws.cell(1, 1, "Complete Fleet Status").font = Font(bold=True, size=14)
    headers = ["Vehicle", "Status", "Speed", "Odometer", "Daily Km", "Overspeed Events", "Max Speed", "Last Update"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(3, c, h)
        cell.font = nav_font
        cell.fill = nav_fill
        cell.border = border
    for r, v in enumerate(vehicles, 4):
        events = overspeed_data.get(v["raw_name"], [])
        max_speed = max((int(e.get("speed", 0)) for e in events), default=0)
        ws.cell(r, 1, v["name"])
        ws.cell(r, 2, v["status"])
        ws.cell(r, 3, v["speed_kmh"])
        ws.cell(r, 4, round(v["odometer_km"], 1))
        ws.cell(r, 5, v["daily_km"])
        ws.cell(r, 6, len(events))
        ws.cell(r, 7, max_speed)
        ws.cell(r, 8, v["last_update"])
        for c in range(1, 9):
            ws.cell(r, c).border = border
    for c in range(1, 9):
        ws.column_dimensions[chr(64 + c)].width = 18
    
    fname = f"MKV_GPS2_ETIQAN_{date_str}.xlsx"
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
    wb.save(out)
    return fname


# ─── MAIN ───────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="YYYY-MM-DD (default: today)")
    parser.add_argument("--dry-run", action="store_true", help="No Slack/file save")
    args = parser.parse_args()
    
    if args.date:
        date_obj = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=DUBAI_OFFSET)
    else:
        date_obj = datetime.now(DUBAI_OFFSET)
    date_str = date_obj.strftime("%Y-%m-%d")
    yesterday_str = (date_obj - timedelta(days=1)).strftime("%Y-%m-%d")
    
    print("\n" + "=" * 60)
    print("  MKV CAR RENTAL - GPS2: AL ETIQAN Report")
    print(f"  Date: {date_str}")
    if args.dry_run:
        print("  MODE: DRY-RUN (no Slack/save)")
    print("=" * 60)
    
    # Step 1: Fetch vehicles
    print("\n[1] Fetching live status...")
    vehicles = fetch_vehicles()
    if not vehicles:
        print("ERROR: No vehicles fetched")
        sys.exit(1)
    
    # Step 2: Load snapshot & calculate deltas
    print("\n[2] Loading snapshot & calculating deltas...")
    snapshot = load_snapshot()
    if snapshot:
        print(f"  Snapshot date: {snapshot.get('date')}")
        calculate_deltas(vehicles, snapshot)
    else:
        print("  No previous snapshot found (first run)")
    
    # Step 3: Fetch overspeed events for yesterday
    print(f"\n[3] Fetching overspeed events for {yesterday_str}...")
    overspeed_data = {}
    for v in vehicles:
        events = fetch_overspeed_events(v["raw_name"], yesterday_str, yesterday_str)
        overspeed_data[v["raw_name"]] = events
        if events:
            print(f"  {v['name']}: {len(events)} overspeed events")
        else:
            print(f"  {v['name']}: no events")
    
    # Step 4: Build Slack messages
    print("\n[4] Building Slack messages...")
    messages = build_slack_messages(vehicles, date_str, overspeed_data)
    
    # Step 5: Post to Slack
    post_to_slack(messages, dry_run=args.dry_run)
    
    # Step 6: Save snapshot & Excel
    if not args.dry_run:
        print("\n[6] Saving snapshot & Excel...")
        save_snapshot(vehicles, date_str)
        fname = build_excel(vehicles, date_str, overspeed_data)
        print(f"  Excel saved: {fname}")
    else:
        print("\n[6] DRY-RUN: snapshot & Excel SKIPPED")
        print("  Would save snapshot with these keys:")
        for v in vehicles:
            print(f"    '{v['device_id']}': {v['odometer_km']}")
    
    print("\n" + "=" * 60)
    print("  ✓ Report complete")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
