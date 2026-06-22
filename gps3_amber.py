"""
gps3_amber.py — MKV CAR RENTAL GPS3: AmberConnect Daily Report
AmberConnect Open API v1 — no PHPSESSID, no Selenium.
Slack format matches GPS1 exactly (4 messages).
"""
import argparse, io, json, os, sys, time
from datetime import datetime, timedelta, timezone
from pathlib import Path

if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True, errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", line_buffering=True, errors="replace")
    except (AttributeError, io.UnsupportedOperation):
        pass

import openpyxl, requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
load_dotenv()

# ── Config ────────────────────────────────────────────────────────────────────
AMBER_API_BASE = "https://api.amberconnect.com/v1/openapi"
AMBER_API_KEY  = os.getenv("AMBER_API_KEY", "A83UAI9QUZC0P8GJ70C16V3RXUVU91P3COV0KDE9J")
DEVICE_OFFSET  = "14400"
SLACK_TOKEN    = os.getenv("SLACK_TOKEN", os.getenv("SLACK_BOT_TOKEN", ""))
SLACK_CHANNEL  = os.getenv("SLACK_CHANNEL", "C0B6Y6EG85D")
TIMEZONE_NAME  = "Asia/Dubai"
DUBAI          = timezone(timedelta(hours=4))
SCRIPT_DIR     = Path(__file__).resolve().parent
SNAPSHOT_FILE  = SCRIPT_DIR / "gps3_amber_snapshot.json"
TIMEOUT        = 30

FLEET = [
    {"name": "Land Rover Defender AA 78043", "token": "GJFZITHNQK84", "imei": "0869066063071345"},
    {"name": "Range Rover Velar AA 68620",   "token": "P7OSKXQ5UAZV", "imei": "0869066063057401"},
    {"name": "Chevrolet Tahoe AA 78042",     "token": "S2Y87JVXG1WB", "imei": "0869066063060660"},
    {"name": "BMW 735i 22 93950",            "token": "ZPB6Y2NREIUX", "imei": "0869066065167372"},
    {"name": "Ford Mustang AA 77491",        "token": "MJ59XPRUY43G", "imei": "0869066063062682"},
    {"name": "Ford Mustang AA 77490",        "token": "3PNTSJ9KLE12", "imei": "0869066063067640"},
    {"name": "Mercedes C200 AA 78067",       "token": "D46GMYAH1ZVF", "imei": "0869066063071212"},
]

SPEED_LIMIT    = 120; URGENT_SPEED = 160; URGENT_KM = 250
WATCH_KM_MIN   = 150; URGENT_OFF_H = 6;  LONG_IDLE_H = 48

# ── Excel colors (match GPS1) ─────────────────────────────────────────────────
DARK_GREEN="1A3C2E"; MID_GREEN="2E6B4F"; WHITE="FFFFFF"
LIGHT_GREY="F5F5F5"; RED="CC0000"; AMBER_COL="CC7700"

def hfont(bold=True,color=WHITE,size=11): return Font(name="Arial",bold=bold,color=color,size=size)
def bfont(bold=False,color="000000",size=10): return Font(name="Arial",bold=bold,color=color,size=size)
def hfill(color=DARK_GREEN): return PatternFill("solid",fgColor=color)
def afill(): return PatternFill("solid",fgColor=LIGHT_GREY)
def bdr():
    s=Side(style="thin",color="CCCCCC"); return Border(left=s,right=s,top=s,bottom=s)
def center(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
def left_a(): return Alignment(horizontal="left",vertical="center",wrap_text=True)
def style_header(ws,row,cols):
    for col in range(1,cols+1):
        c=ws.cell(row=row,column=col); c.font=hfont(); c.fill=hfill(); c.alignment=center(); c.border=bdr()
def style_row(ws,row,cols,alt=False):
    for col in range(1,cols+1):
        c=ws.cell(row=row,column=col); c.font=bfont(); c.border=bdr(); c.alignment=left_a()
        if alt: c.fill=afill()
def title_block(ws,title,date_str,span):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    t=ws.cell(1,1,f"MKV CAR RENTAL \u2014 GPS3: AmberConnect \u2014 {title}")
    t.font=Font(name="Arial",bold=True,size=14,color=WHITE); t.fill=hfill(DARK_GREEN); t.alignment=center(); ws.row_dimensions[1].height=28
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=span)
    d=ws.cell(2,1,f"Report Date: {date_str}  |  Timezone: {TIMEZONE_NAME}  |  Generated: {datetime.now(DUBAI).strftime('%Y-%m-%d %H:%M')}")
    d.font=Font(name="Arial",size=9,color=WHITE); d.fill=hfill(MID_GREEN); d.alignment=center(); ws.row_dimensions[2].height=18
def set_widths(ws,widths):
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

# ── Snapshot ──────────────────────────────────────────────────────────────────
def load_snapshot():
    try:
        if SNAPSHOT_FILE.exists():
            data = json.loads(SNAPSHOT_FILE.read_text(encoding="utf-8"))
            # Ensure all expected keys exist (safe migration from older format)
            data.setdefault("odometers", {})
            data.setdefault("last_move_ts", {})
            data.setdefault("date", "")
            return data
    except Exception:
        pass
    return {"date": "", "odometers": {}, "last_move_ts": {}}

def save_snapshot(vehicles, date_str):
    snap = load_snapshot()
    snap["date"] = date_str
    for v in vehicles:
        imei = v.get("imei", "")
        if not imei:
            continue
        odo = v.get("odometer_km", 0)
        lmt = v.get("last_move_ts")
        if odo and odo > 0:
            snap["odometers"][imei] = odo
        if lmt:
            snap["last_move_ts"][imei] = lmt
    try:
        SNAPSHOT_FILE.write_text(json.dumps(snap), encoding="utf-8")
        print(f"  Snapshot saved ({len(snap['odometers'])} vehicles)")
    except OSError as e:
        print(f"  [warn] snapshot save error: {e}")

# ── API ───────────────────────────────────────────────────────────────────────
def amber_post(endpoint, body):
    try:
        r = requests.post(f"{AMBER_API_BASE}/{endpoint}", json=body, timeout=TIMEOUT)
        return r.status_code, r.json()
    except Exception as e:
        print(f"  [ERROR] {endpoint}: {e}"); return 0, {}

def parse_odometer(raw):
    try:
        val = float(raw or 0)
        if val > 200000: val = round(val / 1000, 1)  # meters -> km
        return round(val, 1)
    except Exception:
        return 0.0

# ── Risk (match GPS1) ─────────────────────────────────────────────────────────
def classify_risk(v):
    reasons = []; tier = "Normal"
    age_h = v.get("age_hours", 0)
    dk    = v.get("daily_km", 0) or 0
    ms    = v.get("max_speed", 0) or 0
    if age_h > URGENT_OFF_H:   reasons.append(f"Offline {age_h:.0f}h"); tier = "Urgent"
    if ms > URGENT_SPEED:      reasons.append(f"Peak {ms} km/h");       tier = "Urgent"
    if dk > URGENT_KM:         reasons.append(f"{dk:.0f} km driven");   tier = "Urgent"
    if tier == "Normal":
        if 0.5 < age_h <= URGENT_OFF_H:     reasons.append(f"Last update {age_h:.1f}h ago"); tier = "Watch"
        if SPEED_LIMIT < ms <= URGENT_SPEED: reasons.append(f"Peak {ms} km/h");              tier = "Watch"
        if WATCH_KM_MIN < dk <= URGENT_KM:  reasons.append(f"{dk:.0f} km driven");           tier = "Watch"
    return tier, reasons

# ── Build records ─────────────────────────────────────────────────────────────
def build_vehicle_records(report_date):
    today_dxb = report_date.astimezone(DUBAI).replace(hour=0, minute=0, second=0, microsecond=0)
    yest_dxb  = today_dxb - timedelta(days=1)
    date_from = yest_dxb.strftime("%Y-%m-%d %H:%M:%S")
    date_to   = today_dxb.strftime("%Y-%m-%d %H:%M:%S")
    now       = time.time()
    snap      = load_snapshot()

    # Batch live — 1 call for all 7
    print("  Fetching live status (batch)...")
    _, body = amber_post("getbulklivetracking", {
        "DeviceOffset": DEVICE_OFFSET,
        "AmberAuthToken": ",".join(v["token"] for v in FLEET),
        "APIKey": AMBER_API_KEY, "DistanceUnit": "kms"})
    live_map = {}
    if body.get("success") == "Y":
        for d in (body.get("Details") or []):
            live_map[str(d.get("IMEI"))] = d
    print(f"  Got live data for {len(live_map)} vehicles")

    records = []
    for fv in FLEET:
        imei = fv["imei"]; name = fv["name"]; token = fv["token"]
        live = live_map.get(imei, {})

        speed        = float(live.get("Speed", 0) or 0)
        ignition     = int(live.get("IgnitionStatus", 0) or 0)
        parking_flag = str(live.get("ParkingFlag", "Y")).upper()
        odometer_km  = parse_odometer(live.get("OdometerReading", 0))
        dt_str       = live.get("DateTime", "")
        last_hb      = live.get("LastHeartBeatTime", "") or dt_str

        age_hours = 0.0
        if last_hb:
            try:
                dt_utc = datetime.strptime(last_hb, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                age_hours = max(0.0, (now - dt_utc.timestamp()) / 3600.0)
            except ValueError:
                pass

        online    = age_hours < 0.5
        is_moving = speed > 0 and parking_flag == "N"

        last_seen = ""
        if dt_str:
            try:
                utc_dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
                last_seen = utc_dt.astimezone(DUBAI).strftime("%Y-%m-%d %H:%M")
            except ValueError:
                pass

        # Per-vehicle trips (yesterday)
        print(f"    Trips: {name}...")
        _, tbody = amber_post("gettrips", {
            "DeviceOffset": DEVICE_OFFSET, "AmberAuthToken": token,
            "APIKey": AMBER_API_KEY, "Filter": "Custom",
            "CustomStartDate": date_from, "CustomEndDate": date_to,
            "DistanceUnit": "kms", "Page": "1", "Limit": "15"})
        trips = (tbody.get("Trips") or []) if tbody.get("success") == "Y" else []

        daily_km = 0.0; max_speed = 0; last_move_ts = None
        for t in trips:
            daily_km += float(t.get("Distance", 0) or 0)
            ts_val = int(t.get("TopSpeed", 0) or 0)
            if ts_val > max_speed: max_speed = ts_val
            end_str = t.get("EndTime", "")
            if end_str:
                try:
                    end_ts = int(datetime.strptime(end_str, "%Y-%m-%d %H:%M:%S")
                                 .replace(tzinfo=timezone.utc).timestamp())
                    if last_move_ts is None or end_ts > last_move_ts:
                        last_move_ts = end_ts
                except ValueError:
                    pass
        daily_km  = round(daily_km, 1)
        trips_cnt = len(trips)

        # Parking duration
        lmt = last_move_ts or snap.get("last_move_ts", {}).get(imei)
        if lmt:
            parking_hrs       = max(0.0, (now - int(lmt)) / 3600.0)
            parking_hrs_known = True
            last_move = datetime.fromtimestamp(int(lmt), tz=DUBAI).strftime("%Y-%m-%d %H:%M")
        else:
            parking_hrs = 0.0; parking_hrs_known = False; last_move = ""

        rec = {
            "name": name, "imei": imei, "token": token,
            "status": "Moving" if is_moving else "Stopped",
            "speed": speed, "ignition": ignition, "odometer_km": odometer_km,
            "daily_km": daily_km, "max_speed": max_speed, "trips_cnt": trips_cnt,
            "last_seen": last_seen, "last_move": last_move, "last_move_ts": last_move_ts,
            "age_hours": age_hours, "online": online, "is_moving": is_moving,
            "moved_today": trips_cnt > 0 and daily_km > 0,
            "parking_hrs": parking_hrs, "parking_hrs_known": parking_hrs_known,
        }
        risk, reasons = classify_risk(rec)
        rec["risk"] = risk; rec["risk_reasons"] = reasons
        records.append(rec)
    return records

# ── Slack helpers ─────────────────────────────────────────────────────────────
def _trunc(s, n):
    s = s or ""
    if len(s) <= n: return s
    cut = s[:n-1]; sp = cut.rfind(" ")
    if sp >= n-9: cut = cut[:sp]
    return cut + "\u2026"

def fmt_idle(hours, known=True):
    if not known or hours is None: return "\u2014"
    if hours < 48: return f"{hours:.1f} hrs"
    return f"{hours/24:.1f} days"

# ── MSG1 — Snapshot (match GPS1 exactly) ─────────────────────────────────────
def build_msg1(vehicles, date_str):
    online   = sum(1 for v in vehicles if v["online"])
    offline  = len(vehicles) - online
    moved    = sum(1 for v in vehicles if v["moved_today"])
    total_km = sum(v["daily_km"] for v in vehicles)
    urgent   = sum(1 for v in vehicles if v["risk"] == "Urgent")
    watch    = sum(1 for v in vehicles if v["risk"] == "Watch")
    return (
        f":car: *MKV CAR RENTAL \u2014 GPS3: AmberConnect Daily Report*\n"
        f"*Date:* {date_str}  |  *Timezone:* {TIMEZONE_NAME}\n"
        f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
        f":white_check_mark: *Online:* {online} / {len(vehicles)}\n"
        f":red_circle: *Offline:* {offline}\n"
        f":chart_with_upwards_trend: *Moved Today:* {moved} vehicles\n"
        f":round_pushpin: *Total Km Driven Today:* {total_km:,.1f} km\n"
        f":rotating_light: *Needs Attention:* {urgent+watch}  ({urgent} Urgent \u00b7 {watch} Watch)\n"
        f":lock: *Engine Blocked:* \u2014\n"
        f"\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\n"
        f"_Generated at {datetime.now(DUBAI).strftime('%Y-%m-%d %H:%M')} (Dubai time)_"
    )

# ── MSG2 — Moved Today (match GPS1 exactly) ───────────────────────────────────
def build_msg2(vehicles):
    moved = sorted([v for v in vehicles if v["moved_today"]], key=lambda x: x["name"].lower())
    lines = [f"{'Vehicle':<38} {'Km Today':>10}   Zone",
             "\u2500" * 72]
    for v in moved:
        lines.append(f"{_trunc(v['name'],37):<38} {v['daily_km']:>9.1f}km   \u2014")
    if not moved:
        lines.append("No vehicles moved today")
    return (
        f":blue_car: *GPS3 \u2014 Vehicles That Moved Today ({len(moved)})*\n"
        f"```\n" + "\n".join(lines) + "\n```"
    )

# ── MSG3 — Parked split <48h / >48h (match GPS1 exactly) ─────────────────────
def build_msg3(vehicles):
    stationary = [v for v in vehicles if not v["moved_today"] and v["online"]]
    offline_v  = [v for v in vehicles if not v["online"]]
    short_idle = sorted([v for v in stationary
                         if not v["parking_hrs_known"] or v["parking_hrs"] < LONG_IDLE_H],
                        key=lambda x: x["parking_hrs"], reverse=True)
    long_idle  = sorted([v for v in stationary
                         if v["parking_hrs_known"] and v["parking_hrs"] >= LONG_IDLE_H],
                        key=lambda x: x["parking_hrs"], reverse=True)

    def _tbl(rows):
        lines = [f"{'Vehicle':<38} {'Idle Time':>11}   Zone",
                 "\u2500" * 72]
        for v in rows:
            dur = fmt_idle(v["parking_hrs"], v["parking_hrs_known"])
            lines.append(f"{_trunc(v['name'],37):<38} {dur:>11}   \u2014")
        return "\n".join(lines)

    msg  = (f":parking: *GPS3 \u2014 Idle <48h ({len(short_idle)} vehicles)*\n"
            f"```\n{_tbl(short_idle)}\n```\n")
    msg += (f"\n:package: *GPS3 \u2014 Long-idle >48h ({len(long_idle)} vehicles)*"
            f" \u2014 review for utilization\n"
            f"```\n{_tbl(long_idle)}\n```")
    if offline_v:
        msg += (f"\n_:red_circle: {len(offline_v)} vehicle"
                f"{'s' if len(offline_v)!=1 else ''} offline \u2014 excluded from list._")
    return msg

# ── MSG4 — Speeding deduped (match GPS1 exactly) ──────────────────────────────
def build_msg4(vehicles):
    offenders = sorted([v for v in vehicles if (v["max_speed"] or 0) > SPEED_LIMIT],
                       key=lambda x: x["max_speed"], reverse=True)
    if not offenders:
        return (f":white_check_mark: *GPS3 \u2014 No speeding events recorded yesterday"
                f" (>{SPEED_LIMIT} km/h)*")
    lines = [f"{'Vehicle':<38} {'Peak':>5}  {'Avg':>5}  Trips",
             "\u2500" * 65]
    for v in offenders:
        lines.append(f"{_trunc(v['name'],37):<38} {v['max_speed']:>5}      \u2014  {v['trips_cnt']:>5}")
    return (
        f":rotating_light: *GPS3 \u2014 Speeding Yesterday \u2014 "
        f"{len(offenders)} vehicle{'s' if len(offenders)!=1 else ''} over {SPEED_LIMIT} km/h*\n"
        f"```\n" + "\n".join(lines) + "\n```"
    )

def post_slack(messages, dry_run=False):
    if dry_run or not SLACK_TOKEN:
        for i, m in enumerate(messages, 1):
            print(f"\n\u2500\u2500\u2500\u2500 MSG{i} \u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500\u2500")
            print(m)
        if not SLACK_TOKEN: print("\n  [WARN] No SLACK_TOKEN \u2014 not posting")
        return
    hdrs = {"Authorization": f"Bearer {SLACK_TOKEN}", "Content-Type": "application/json"}
    ok = 0
    for m in messages:
        try:
            r = requests.post("https://slack.com/api/chat.postMessage",
                              headers=hdrs, json={"channel": SLACK_CHANNEL, "text": m}, timeout=10)
            if r.json().get("ok"): ok += 1
            else: print(f"  [Slack] Error: {r.json().get('error')}")
        except Exception as e: print(f"  [Slack] Exception: {e}")
    print(f"  [Slack] Posted {ok}/4 messages to #daily-gps-update")

# ── Excel ─────────────────────────────────────────────────────────────────────
def build_summary(wb, vehicles, date_str):
    ws = wb.create_sheet("GPS3 Summary", 0); ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=30; ws.column_dimensions["C"].width=25
    ws.merge_cells("B1:C1"); t=ws["B1"]
    t.value="MKV CAR RENTAL \u2014 GPS3: AmberConnect \u2014 Daily Report"
    t.font=Font(name="Arial",bold=True,size=16,color=WHITE); t.fill=hfill(DARK_GREEN); t.alignment=center(); ws.row_dimensions[1].height=36
    ws.merge_cells("B2:C2"); s=ws["B2"]
    s.value=f"{date_str}  |  {TIMEZONE_NAME}  |  {datetime.now(DUBAI).strftime('%Y-%m-%d %H:%M')}"
    s.font=Font(name="Arial",size=10,color=WHITE); s.fill=hfill(MID_GREEN); s.alignment=center(); ws.row_dimensions[2].height=20
    ws["B4"]="Fleet Summary \u2014 GPS3: AmberConnect"; ws["B4"].font=Font(name="Arial",bold=True,size=12,color=DARK_GREEN)
    stats=[("Total Vehicles",len(vehicles)),("Online",sum(1 for v in vehicles if v["online"])),
           ("Offline",sum(1 for v in vehicles if not v["online"])),
           ("Moved Today",f"{sum(1 for v in vehicles if v['moved_today'])} vehicles"),
           ("Urgent",sum(1 for v in vehicles if v["risk"]=="Urgent")),
           ("Watch",sum(1 for v in vehicles if v["risk"]=="Watch")),
           ("Km Driven Today",f"{sum(v['daily_km'] for v in vehicles):,.1f} km"),
           ("Data Source","AmberConnect Open API v1"),
           ("Generated At",datetime.now(DUBAI).strftime("%Y-%m-%d %H:%M"))]
    for row_i,(label,val) in enumerate(stats,5):
        ws.cell(row_i,2).value=label; ws.cell(row_i,2).font=Font(name="Arial",bold=True,size=10)
        ws.cell(row_i,2).fill=PatternFill("solid",fgColor=LIGHT_GREY); ws.cell(row_i,3).value=str(val)
        ws.cell(row_i,3).font=Font(name="Arial",size=10); ws.row_dimensions[row_i].height=18

def build_daily_movement(wb, vehicles, date_str):
    ws=wb.create_sheet("GPS3 Daily Movement")
    headers=["#","Vehicle Name","Moved Today","Km Yesterday","Max Speed","Trips","Last Move","Status","Online","Risk"]
    widths=[5,35,13,14,11,7,20,12,10,10]
    title_block(ws,"DAILY MOVEMENT REPORT",date_str,len(headers))
    ws.append([]); ws.append(headers); style_header(ws,4,len(headers)); set_widths(ws,widths)
    ds=5
    for i,v in enumerate(vehicles,1):
        status="Moving" if v["is_moving"] else "Stopped" if v["online"] else "Offline"
        ws.append([i,v["name"],"\u2713 Yes" if v["moved_today"] else "No",v["daily_km"],v["max_speed"],v["trips_cnt"],v["last_move"],status,"Online" if v["online"] else "Offline",v["risk"]])
        style_row(ws,ds+i-1,len(headers),alt=(i%2==0))
        rc=ws.cell(ds+i-1,10)
        if v["risk"]=="Urgent": rc.font=Font(name="Arial",bold=True,size=10,color=RED)
        elif v["risk"]=="Watch": rc.font=Font(name="Arial",bold=True,size=10,color=AMBER_COL)
    ws.freeze_panes="A5"

def build_parked(wb, vehicles, date_str):
    ws=wb.create_sheet("GPS3 Parked Vehicles")
    headers=["#","Vehicle Name","Status","Idle Time","Last Move","Online","Last Seen"]
    widths=[5,35,12,14,20,10,20]
    title_block(ws,"PARKED VEHICLES REPORT",date_str,len(headers))
    ws.append([]); ws.append(headers); style_header(ws,4,len(headers)); set_widths(ws,widths)
    parked=sorted([v for v in vehicles if not v["moved_today"]],key=lambda x:x["parking_hrs"],reverse=True)
    ds=5
    for i,v in enumerate(parked,1):
        ws.append([i,v["name"],"Moving" if v["is_moving"] else "Stopped",fmt_idle(v["parking_hrs"],v["parking_hrs_known"]),v["last_move"] or "\u2014","Online" if v["online"] else "Offline",v["last_seen"]])
        style_row(ws,ds+i-1,len(headers),alt=(i%2==0))
    ws.freeze_panes="A5"

def build_fleet(wb, vehicles, date_str):
    ws=wb.create_sheet("GPS3 Fleet Status")
    headers=["#","Vehicle Name","IMEI","Status","Speed (km/h)","Ignition","Km Today","Odometer (km)","Last Seen","Risk"]
    widths=[5,35,18,12,13,10,12,14,20,10]
    title_block(ws,"FULL FLEET STATUS",date_str,len(headers))
    ws.append([]); ws.append(headers); style_header(ws,4,len(headers)); set_widths(ws,widths)
    ds=5
    for i,v in enumerate(vehicles,1):
        ws.append([i,v["name"],v["imei"],"Moving" if v["is_moving"] else "Stopped",v["speed"],"ON" if v["ignition"] else "off",v["daily_km"],v["odometer_km"],v["last_seen"],v["risk"]])
        style_row(ws,ds+i-1,len(headers),alt=(i%2==0))
        rc=ws.cell(ds+i-1,10)
        if v["risk"]=="Urgent": rc.font=Font(name="Arial",bold=True,size=10,color=RED)
        elif v["risk"]=="Watch": rc.font=Font(name="Arial",bold=True,size=10,color=AMBER_COL)
    ws.freeze_panes="A5"

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser=argparse.ArgumentParser()
    parser.add_argument("--date",help="Report date YYYY-MM-DD")
    parser.add_argument("--dry-run",action="store_true")
    parser.add_argument("--test",action="store_true")
    args=parser.parse_args()

    report_date=datetime.now(DUBAI)
    if args.date: report_date=datetime.strptime(args.date,"%Y-%m-%d").replace(tzinfo=DUBAI)
    date_str=report_date.strftime("%Y-%m-%d")

    print(f"\n{'='*55}\n  MKV CAR RENTAL - GPS3: AmberConnect Report\n  Date: {date_str}")
    if args.dry_run: print("  MODE: dry-run (no Slack post)")
    print(f"{'='*55}\n")

    if args.test:
        vehicles=[
            {"name":"Land Rover Defender AA 78043","imei":"0869066063071345","token":"GJFZITHNQK84","status":"Stopped","speed":0,"ignition":0,"odometer_km":51.7,"daily_km":45.2,"max_speed":0,"trips_cnt":2,"last_seen":"2026-06-22 10:30","last_move":"2026-06-21 18:00","last_move_ts":None,"age_hours":0.3,"online":True,"is_moving":False,"moved_today":True,"parking_hrs":16.5,"parking_hrs_known":True,"risk":"Normal","risk_reasons":[]},
            {"name":"Mercedes C200 AA 78067","imei":"0869066063071212","token":"D46GMYAH1ZVF","status":"Stopped","speed":0,"ignition":0,"odometer_km":43.4,"daily_km":207.2,"max_speed":138,"trips_cnt":5,"last_seen":"2026-06-22 09:00","last_move":"2026-06-21 20:00","last_move_ts":None,"age_hours":0.2,"online":True,"is_moving":False,"moved_today":True,"parking_hrs":14.0,"parking_hrs_known":True,"risk":"Watch","risk_reasons":["207 km driven"]},
        ]
    else:
        vehicles=build_vehicle_records(report_date)

    online=sum(1 for v in vehicles if v["online"]); moved=sum(1 for v in vehicles if v["moved_today"])
    total_km=sum(v["daily_km"] for v in vehicles)
    urgent=sum(1 for v in vehicles if v["risk"]=="Urgent"); watch=sum(1 for v in vehicles if v["risk"]=="Watch")
    print(f"  Vehicles: {len(vehicles)} | Online: {online} | Moved: {moved} | Total km: {total_km:,.1f}")
    print(f"  Urgent: {urgent} | Watch: {watch}")
    for v in vehicles:
        flag=f"[{v['risk']}]" if v["risk"]!="Normal" else ""
        print(f"  {v['name']:<40} {v['daily_km']:>7.1f} km  {v['status']:<8} {flag}")

    print("\nBuilding Excel...")
    wb=openpyxl.Workbook(); wb.remove(wb.active)
    build_summary(wb,vehicles,date_str); build_daily_movement(wb,vehicles,date_str)
    build_parked(wb,vehicles,date_str); build_fleet(wb,vehicles,date_str)
    fname=f"MKV_GPS3_AMBER_{date_str}.xlsx"; wb.save(SCRIPT_DIR/fname); print(f"  Saved: {fname}")

    if not args.test:
        save_snapshot(vehicles,date_str)

    post_slack([build_msg1(vehicles,date_str),build_msg2(vehicles),build_msg3(vehicles),build_msg4(vehicles)],dry_run=args.dry_run)
    print(f"\n{'='*55}\n")

if __name__=="__main__":
    main()
