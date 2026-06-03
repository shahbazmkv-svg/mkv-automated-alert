"""
MKV CAR RENTAL — Pilot GPS Daily Report Generator
Generates per-vehicle daily movement report for all 64 vehicles.
Stores odometer snapshot daily to calculate km driven per vehicle.

Usage:
    python daily_gps_report.py
    python daily_gps_report.py --date 2026-06-02

Requirements:
    pip install requests openpyxl python-dotenv
"""

import requests
import openpyxl
import json
import os
import sys
import argparse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta, timezone
from dotenv import load_dotenv

load_dotenv()

API_LOGIN_URL    = "https://pilot-gps.ru/backend/ax/user/login.php"
API_CURRENT_DATA = "https://pilot-gps.ru/backend/ax/current_data.php"
API_REPORTS_URL  = "https://pilot-gps.ru/backend/ax/reports.php"
API_USER         = os.getenv("PILOT_USER", "")
API_PASS         = os.getenv("PILOT_PASS", "")
SLACK_TOKEN      = os.getenv("SLACK_TOKEN", "")
SLACK_CHANNEL    = os.getenv("SLACK_CHANNEL", "C0B6Y6EG85D")
TIMEZONE         = "Asia/Dubai"
DUBAI_OFFSET     = timezone(timedelta(hours=4))

DARK_GREEN = "1A3C2E"
MID_GREEN  = "2E6B4F"
WHITE      = "FFFFFF"
LIGHT_GREY = "F5F5F5"
RED        = "CC0000"
AMBER      = "CC7700"
GREY       = "999999"
GREEN      = "2E6B4F"

def hfont(bold=True, color=WHITE, size=11):
    return Font(name="Arial", bold=bold, color=color, size=size)
def bfont(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)
def hfill(color=DARK_GREEN):
    return PatternFill("solid", fgColor=color)
def afill():
    return PatternFill("solid", fgColor=LIGHT_GREY)
def bdr():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = hfont(); c.fill = hfill(); c.alignment = center(); c.border = bdr()

def style_row(ws, row, cols, alt=False):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.font = bfont(); c.border = bdr(); c.alignment = left()
        if alt: c.fill = afill()

def title_block(ws, title, date_str, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    t = ws.cell(1, 1, f"MKV CAR RENTAL — {title}")
    t.font = Font(name="Arial", bold=True, size=14, color=WHITE)
    t.fill = hfill(DARK_GREEN); t.alignment = center()
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    d = ws.cell(2, 1, f"Report Date: {date_str}  |  Timezone: {TIMEZONE}  |  Generated: {datetime.now(DUBAI_OFFSET).strftime('%Y-%m-%d %H:%M')}")
    d.font = Font(name="Arial", size=9, color=WHITE)
    d.fill = hfill(MID_GREEN); d.alignment = center()
    ws.row_dimensions[2].height = 18

def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def totals_row(ws, row, ncols, formulas, label_col=1):
    ws.cell(row, label_col).value = "TOTAL / AVERAGE"
    ws.cell(row, label_col).font = Font(name="Arial", bold=True, color=WHITE)
    ws.cell(row, label_col).fill = hfill(MID_GREEN)
    ws.cell(row, label_col).alignment = center()
    for col, formula in formulas.items():
        c = ws.cell(row, col)
        c.value = formula
        c.font = Font(name="Arial", bold=True, color=WHITE)
        c.fill = hfill(MID_GREEN)
        c.alignment = center(); c.border = bdr()
    for col in range(1, ncols + 1):
        ws.cell(row, col).border = bdr()


# ── SNAPSHOT (for daily km calculation) ──────────────────────────────────────
def load_snapshot():
    try:
        if os.path.exists(SNAPSHOT_FILE):
            with open(SNAPSHOT_FILE) as f:
                return json.load(f)
    except: pass
    return {}

def save_snapshot(vehicles, date_str):
    snap = {
        "date": date_str,
        "odometers": {str(v["id"]): v["total_km"] for v in vehicles}
    }
    with open(SNAPSHOT_FILE, "w") as f:
        json.dump(snap, f)

def calc_daily_km(vehicle, snapshot):
    """Calculate km driven today using odometer delta."""
    today = datetime.now(DUBAI_OFFSET).strftime("%Y-%m-%d")
    snap_date = snapshot.get("date", "")
    snap_odos = snapshot.get("odometers", {})
    vid = str(vehicle["id"])

    # Only use snapshot if it's from today or yesterday
    if snap_date in [today, (datetime.now(DUBAI_OFFSET) - timedelta(days=1)).strftime("%Y-%m-%d")]:
        prev_km = snap_odos.get(vid)
        if prev_km is not None:
            delta = round(vehicle["total_km"] - prev_km, 1)
            return max(0, delta)  # no negative

    # Fallback: estimate from last_move timestamp
    last_move_ts = vehicle.get("last_move_ts")
    if last_move_ts:
        try:
            today_start = int(datetime.now(DUBAI_OFFSET).replace(
                hour=0, minute=0, second=0, microsecond=0).timestamp())
            if int(last_move_ts) >= today_start:
                return "Moved today"
        except: pass
    return "-"


# ── API CLIENT ────────────────────────────────────────────────────────────────
class PilotAPI:
    def __init__(self):
        self.session = requests.Session()
        print(f"  Auth: logging in as {API_USER}")
        r = self.session.post(API_LOGIN_URL,
                              data={"username": API_USER, "password": API_PASS}, timeout=15)
        data = r.json()
        if r.status_code == 200 and data.get("success"):
            print("  Auth: login successful ✓")
        else:
            print("  Auth: login failed"); sys.exit(1)

    def get_daily_mileage(self, vehicle_ids, report_date):
        """Fetch yesterday's km driven per vehicle using reports API."""
        yesterday = report_date - timedelta(days=1)
        start_date  = yesterday.strftime("%d.%m.%Y 00:00")
        stop_date   = report_date.strftime("%d.%m.%Y 00:00")
        start_month = yesterday.strftime("%m.%Y")
        stop_month  = report_date.strftime("%m.%Y")

        mileage = {}
        for vid in vehicle_ids:
            try:
                data = {
                    'download': 0, 'start_time': '00:00', 'stop_time': '00:00',
                    'veh_id': vid, 'zones_id': '', 'lines_id': '',
                    'stopping_points_id': '', 'drivers_id': '', 'groups_id': '',
                    'holidays': '', 'lang': 'en', 'explode': 1,
                    'start_month': start_month, 'stop_month': stop_month,
                    'pre_start_date': (yesterday - timedelta(days=7)).strftime("%d.%m.%Y"),
                    'pre_stop_date': yesterday.strftime("%d.%m.%Y"),
                    'start_date': start_date, 'stop_date': stop_date,
                    'report_type': 6,
                }
                r = self.session.post(API_REPORTS_URL, data=data, timeout=15)
                resp = r.json()
                km = 0
                for vname_key, date_data in resp.get('data', {}).items():
                    for date_key, arr in date_data.items():
                        if isinstance(arr, list) and len(arr) > 6:
                            km = float(arr[6] or 0)
                mileage[vid] = round(km, 1)
            except Exception as e:
                mileage[vid] = 0
        return mileage

    def get_speed_violations(self, vehicles, report_date, speed_limit=120):
        """Fetch speeding events per vehicle from yesterday's trips."""
        yesterday = report_date - timedelta(days=1)
        start_date  = yesterday.strftime("%d.%m.%Y 00:00")
        stop_date   = report_date.strftime("%d.%m.%Y 00:00")
        start_month = yesterday.strftime("%m.%Y")
        stop_month  = report_date.strftime("%m.%Y")

        violations = []
        for v in vehicles:
            vid = v["id"]
            vname = v["name"]
            try:
                data = {
                    'download': 0, 'start_time': '00:00', 'stop_time': '00:00',
                    'veh_id': vid, 'zones_id': '', 'lines_id': '',
                    'stopping_points_id': '', 'drivers_id': '', 'groups_id': '',
                    'holidays': '', 'lang': 'en', 'explode': 1,
                    'avg_max_speed': 1,
                    'start_month': start_month, 'stop_month': stop_month,
                    'pre_start_date': (yesterday - timedelta(days=7)).strftime("%d.%m.%Y"),
                    'pre_stop_date': yesterday.strftime("%d.%m.%Y"),
                    'start_date': start_date, 'stop_date': stop_date,
                    'report_type': 1,
                }
                r = self.session.post(API_REPORTS_URL, data=data, timeout=15)
                resp = r.json()
                for vname_key, date_data in resp.get('data', {}).items():
                    for date_range, events in date_data.items():
                        for ts, event in events.items():
                            if event.get('type') == 'Movement':
                                max_spd = event.get('max_speed', 0) or 0
                                if float(max_spd) > speed_limit:
                                    violations.append({
                                        'vehicle': vname,
                                        'max_speed': float(max_spd),
                                        'avg_speed': event.get('avg_speed', 0),
                                        'mileage': event.get('mileage', 0),
                                        'duration_min': round(event.get('duration', 0) / 60, 1),
                                    })
            except Exception:
                pass
        return violations

    def get_vehicles(self):
        r = self.session.get(API_CURRENT_DATA, timeout=30)
        data = r.json()
        vehicles = []
        today_start = int(datetime.now(DUBAI_OFFSET).replace(
            hour=0, minute=0, second=0, microsecond=0).timestamp())

        for obj in data.get("objects", []):
            last_event = obj.get("last_event", {})
            is_moving  = last_event.get("type") == "move"
            speed      = float(last_event.get("speed", 0) or 0)
            last_ts    = obj.get("unixtimestamp")
            last_move  = last_event.get("last_move")
            last_stop  = last_event.get("last_stop")

            def fmt_ts(ts):
                try: return datetime.fromtimestamp(int(ts), tz=DUBAI_OFFSET).strftime("%Y-%m-%d %H:%M") if ts else ""
                except: return ""

            # Did vehicle move today?
            moved_today = False
            if last_move:
                try: moved_today = int(last_move) >= today_start
                except: pass

            # Parking duration
            parking_secs = 0
            event_text = last_event.get("text", "")
            if "|" in event_text:
                try: parking_secs = int(event_text.split("|")[1])
                except: pass

            # Battery & sensors
            sensors = obj.get("sensors", {})
            car_batt = ""
            if "Car Batt" in sensors:
                car_batt = sensors["Car Batt"].split("|")[0]

            engine_blocked = any(
                "Engine block" in k and "Blocked|" in v
                for k, v in sensors.items()
            )

            ignition_on = any(
                "Ignition" in k and "|1|" in v
                for k, v in sensors.items()
            )

            total_km = round(float(obj.get("len", 0) or 0) / 1000, 1)

            vehicles.append({
                "id"            : obj.get("id"),
                "name"          : obj.get("name", "Unknown").strip(),
                "online"        : obj.get("is_server_online", False),
                "is_moving"     : is_moving,
                "moved_today"   : moved_today,
                "speed"         : speed,
                "last_seen"     : fmt_ts(last_ts),
                "last_move"     : fmt_ts(last_move),
                "last_stop"     : fmt_ts(last_stop),
                "last_move_ts"  : last_move,
                "parking_secs"  : parking_secs,
                "parking_hrs"   : round(parking_secs / 3600, 1),
                "zone"          : ", ".join(obj.get("zone", [])) if obj.get("zone") else "",
                "total_km"      : total_km,
                "car_batt"      : car_batt,
                "engine_blocked": engine_blocked,
                "ignition_on"   : ignition_on,
                "event_type"    : last_event.get("type", ""),
                "lat"           : obj.get("lat", ""),
                "lon"           : obj.get("lon", ""),
            })

        print(f"  Found {len(vehicles)} vehicles")
        return vehicles


# ── REPORT BUILDERS ───────────────────────────────────────────────────────────

def build_summary(wb, date_str, vehicles):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 25

    ws.merge_cells("B1:C1")
    t = ws["B1"]
    t.value = "MKV CAR RENTAL — Daily GPS Report"
    t.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    t.fill = hfill(DARK_GREEN); t.alignment = center()
    ws.row_dimensions[1].height = 36

    ws.merge_cells("B2:C2")
    s = ws["B2"]
    s.value = f"{date_str}  |  {TIMEZONE}  |  {datetime.now(DUBAI_OFFSET).strftime('%Y-%m-%d %H:%M')}"
    s.font = Font(name="Arial", size=10, color=WHITE)
    s.fill = hfill(MID_GREEN); s.alignment = center()
    ws.row_dimensions[2].height = 20

    online   = sum(1 for v in vehicles if v["online"])
    offline  = len(vehicles) - online
    moving   = sum(1 for v in vehicles if v["is_moving"])
    moved    = sum(1 for v in vehicles if v["moved_today"])
    parked   = sum(1 for v in vehicles if not v["is_moving"] and v["online"])
    blocked  = sum(1 for v in vehicles if v["engine_blocked"])
    ignition = sum(1 for v in vehicles if v["ignition_on"])
    in_zone  = sum(1 for v in vehicles if v["zone"])
    total_km     = sum(v["total_km"] for v in vehicles)
    total_daily  = sum(v.get("daily_km", 0) for v in vehicles)

    ws["B4"] = "Fleet Summary"
    ws["B4"].font = Font(name="Arial", bold=True, size=12, color=DARK_GREEN)

    stats = [
        ("Total Vehicles",        len(vehicles)),
        ("Online Now",            f"{online} ✓"),
        ("Offline",               offline),
        ("Currently Moving",      moving),
        ("Moved Today",           f"{moved} vehicles"),
        ("Parked (Online)",       parked),
        ("Ignition On",           ignition),
        ("Engine Blocked",        blocked),
        ("In a Geofence Zone",    f"{in_zone} vehicles"),
        ("Km Driven Yesterday",   f"{total_daily:,.1f} km"),
        ("Total Fleet Odometer",  f"{total_km:,.1f} km"),
    ]
    for row_i, (label, val) in enumerate(stats, 5):
        ws.cell(row_i, 2).value = label
        ws.cell(row_i, 2).font = Font(name="Arial", bold=True, size=10)
        ws.cell(row_i, 2).fill = PatternFill("solid", fgColor=LIGHT_GREY)
        ws.cell(row_i, 3).value = str(val)
        ws.cell(row_i, 3).font = Font(name="Arial", size=10)
        ws.row_dimensions[row_i].height = 18

    ws["B17"] = "Report Tabs"
    ws["B17"].font = Font(name="Arial", bold=True, size=12, color=DARK_GREEN)
    tabs = [
        ("Daily Movement",    "Km driven today, last move/stop, moved today flag"),
        ("Activity Report",   "Status, ignition, parking duration, zone"),
        ("Speed & Status",    "Speed, battery, engine block, online/offline"),
        ("Geofence",          "Current zone per vehicle"),
    ]
    for row_i, (tab, desc) in enumerate(tabs, 18):
        ws.cell(row_i, 2).value = tab
        ws.cell(row_i, 2).font = Font(name="Arial", bold=True, size=10, color=MID_GREEN)
        ws.cell(row_i, 3).value = desc
        ws.row_dimensions[row_i].height = 18


def build_daily_movement(wb, vehicles, date_str):
    ws = wb.create_sheet("Daily Movement")
    headers = ["#", "Vehicle Name", "Moved Today", "Km Yesterday",
               "Last Move Time", "Last Stop Time", "Current Status",
               "Online", "Current Zone"]
    widths  = [5, 32, 13, 14, 20, 20, 14, 10, 30]
    title_block(ws, "DAILY MOVEMENT REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        daily_km = v.get("daily_km", 0)
        moved    = "✓ Yes" if v["moved_today"] else "No"
        status   = "Moving" if v["is_moving"] else "Parked" if v["online"] else "Offline"
        online   = "Online" if v["online"] else "Offline"

        ws.append([i, v["name"], moved, daily_km,
                   v["last_move"], v["last_stop"],
                   status, online, v["zone"]])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))

        # Color moved today
        mc = ws.cell(data_start + i - 1, 3)
        mc.font = Font(name="Arial", bold=True, size=10,
                       color=GREEN if v["moved_today"] else GREY)

        # Color km driven
        kc = ws.cell(data_start + i - 1, 4)
        if isinstance(daily_km, (int, float)) and daily_km > 0:
            kc.font = Font(name="Arial", bold=True, size=10, color=GREEN)
        # Color status
        sc = ws.cell(data_start + i - 1, 7)
        sc.font = Font(name="Arial", bold=True, size=10,
                       color=GREEN if v["is_moving"] else (GREY if not v["online"] else AMBER))

        # Color online
        oc = ws.cell(data_start + i - 1, 8)
        oc.font = Font(name="Arial", bold=True, size=10,
                       color=GREEN if v["online"] else GREY)

    # Totals
    tr = data_start + len(vehicles)
    moved_count = sum(1 for v in vehicles if v["moved_today"])
    ws.cell(tr, 1).value = "TOTAL"
    ws.cell(tr, 1).font = Font(name="Arial", bold=True, color=WHITE)
    ws.cell(tr, 1).fill = hfill(MID_GREEN)
    ws.cell(tr, 3).value = f"{moved_count} vehicles moved"
    ws.cell(tr, 3).font = Font(name="Arial", bold=True, color=WHITE)
    ws.cell(tr, 3).fill = hfill(MID_GREEN)
    for col in range(1, len(headers) + 1):
        ws.cell(tr, col).border = bdr()
    ws.freeze_panes = "A5"


def build_activity(wb, vehicles, date_str):
    ws = wb.create_sheet("Activity Report")
    headers = ["#", "Vehicle Name", "Status", "Ignition",
               "Parking Duration (hrs)", "Last Seen", "Zone", "Online"]
    widths  = [5, 32, 12, 10, 22, 20, 30, 10]
    title_block(ws, "ACTIVITY REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        status   = "Moving" if v["is_moving"] else "Parked" if v["online"] else "Offline"
        ignition = "ON" if v["ignition_on"] else "off"
        p_hrs    = v["parking_hrs"] if not v["is_moving"] else "-"
        online   = "Online" if v["online"] else "Offline"

        ws.append([i, v["name"], status, ignition, p_hrs,
                   v["last_seen"], v["zone"], online])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))

        sc = ws.cell(data_start + i - 1, 3)
        sc.font = Font(name="Arial", bold=True, size=10,
                       color=GREEN if v["is_moving"] else (GREY if not v["online"] else AMBER))

        ic = ws.cell(data_start + i - 1, 4)
        ic.font = Font(name="Arial", bold=True, size=10,
                       color=GREEN if v["ignition_on"] else GREY)

        oc = ws.cell(data_start + i - 1, 8)
        oc.font = Font(name="Arial", bold=True, size=10,
                       color=GREEN if v["online"] else GREY)
    ws.freeze_panes = "A5"


def build_speed_status(wb, vehicles, date_str):
    ws = wb.create_sheet("Speed & Status")
    headers = ["#", "Vehicle Name", "Online", "Speed (km/h)",
               "Engine Block", "Battery", "Ignition", "Last Seen"]
    widths  = [5, 32, 10, 14, 16, 16, 10, 20]
    title_block(ws, "SPEED & STATUS REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        online  = "Online" if v["online"] else "Offline"
        blocked = "BLOCKED" if v["engine_blocked"] else "OK"
        ign     = "ON" if v["ignition_on"] else "off"

        ws.append([i, v["name"], online, v["speed"],
                   blocked, v["car_batt"], ign, v["last_seen"]])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))

        ws.cell(data_start+i-1, 3).font = Font(name="Arial", bold=True, size=10,
                                                color=GREEN if v["online"] else GREY)
        if v["speed"] > 120:
            ws.cell(data_start+i-1, 4).font = Font(name="Arial", bold=True, color=RED)
        ws.cell(data_start+i-1, 5).font = Font(name="Arial", bold=True, size=10,
                                                color=RED if v["engine_blocked"] else GREEN)
        ws.cell(data_start+i-1, 7).font = Font(name="Arial", bold=True, size=10,
                                                color=GREEN if v["ignition_on"] else GREY)
    ws.freeze_panes = "A5"


def build_geofence(wb, vehicles, date_str):
    ws = wb.create_sheet("Geofence")
    headers = ["#", "Vehicle Name", "Current Zone", "Status", "Online", "Last Seen"]
    widths  = [5, 32, 40, 12, 10, 20]
    title_block(ws, "GEOFENCE REPORT", date_str, len(headers))
    ws.append([]); ws.append(headers)
    style_header(ws, 4, len(headers))
    set_widths(ws, widths)

    data_start = 5
    for i, v in enumerate(vehicles, 1):
        zone   = v["zone"] or "Outside all zones"
        status = "Moving" if v["is_moving"] else "Parked"
        online = "Online" if v["online"] else "Offline"

        ws.append([i, v["name"], zone, status, online, v["last_seen"]])
        style_row(ws, data_start + i - 1, len(headers), alt=(i % 2 == 0))

        if not v["zone"]:
            ws.cell(data_start+i-1, 3).font = Font(name="Arial", italic=True, color=GREY, size=10)
        ws.cell(data_start+i-1, 5).font = Font(name="Arial", bold=True, size=10,
                                                color=GREEN if v["online"] else GREY)

    in_zone  = sum(1 for v in vehicles if v["zone"])
    out_zone = len(vehicles) - in_zone
    tr = data_start + len(vehicles) + 1
    ws.cell(tr, 2).value = f"In a zone: {in_zone}  |  Outside all zones: {out_zone}"
    ws.cell(tr, 2).font = Font(name="Arial", bold=True, size=10, color=MID_GREEN)
    ws.freeze_panes = "A5"


# ── SLACK ─────────────────────────────────────────────────────────────────────
def post_slack(date_str, vehicles, speed_violations=None):
    if not SLACK_TOKEN:
        print("  [Slack] No token — skipping"); return

    online   = sum(1 for v in vehicles if v["online"])
    moving   = sum(1 for v in vehicles if v["is_moving"])
    moved    = sum(1 for v in vehicles if v["moved_today"])
    offline  = len(vehicles) - online
    blocked  = sum(1 for v in vehicles if v["engine_blocked"])
    ignition = sum(1 for v in vehicles if v["ignition_on"])
    in_zone  = sum(1 for v in vehicles if v["zone"])
    total_daily_km = sum(v.get("daily_km", 0) for v in vehicles)

    # Moving vehicles table
    moved_vehicles = [v for v in vehicles if v["moved_today"]]
    moving_table = "```\n"
    moving_table += f"{'Vehicle':<35} {'Km Today':>10} {'Zone':<25}\n"
    moving_table += "─" * 72 + "\n"
    for v in sorted(moved_vehicles, key=lambda x: x["name"]):
        dk = v.get("daily_km", 0)
        km_str = f"{dk:.1f} km" if dk > 0 else "—"
        zone = v["zone"][:24] if v["zone"] else "—"
        moving_table += f"{v['name'][:34]:<35} {km_str:>10} {zone:<25}\n"
    moving_table += "```"

    # Non-moving vehicles — online/parked only, exclude offline
    stationary = [v for v in vehicles if not v["moved_today"] and v["online"]]
    offline_vehicles = [v for v in vehicles if not v["online"]]
    stationary_table = "```\n"
    stationary_table += f"{'Vehicle':<35} {'Parked (hrs)':>12} {'Zone':<25}\n"
    stationary_table += "─" * 72 + "\n"
    for v in sorted(stationary, key=lambda x: x["parking_hrs"], reverse=True):
        p_hrs = f"{v['parking_hrs']} hrs"
        zone = v["zone"][:24] if v["zone"] else "—"
        stationary_table += f"{v['name'][:34]:<35} {p_hrs:>12} {zone:<25}\n"
    stationary_table += "```"

    msg = (
        f":car: *MKV CAR RENTAL — Daily GPS Report*\n"
        f"*Date:* {date_str}  |  *Timezone:* {TIMEZONE}\n"
        f"─────────────────────────────\n"
        f":white_check_mark: *Online:* {online} / {len(vehicles)}\n"
        f":red_circle: *Offline:* {offline}\n"
        f":blue_car: *Moving Now:* {moving}\n"
        f":chart_with_upwards_trend: *Moved Today:* {moved} vehicles\n"
        f":round_pushpin: *Total Km Driven Today:* {total_daily_km:,.1f} km\n"
        f":busts_in_silhouette: *In a Zone:* {in_zone}\n"
        f":lock: *Engine Blocked:* {blocked}\n"
        f"─────────────────────────────\n"
        f"_Generated at {datetime.now(DUBAI_OFFSET).strftime('%Y-%m-%d %H:%M')} (Dubai time)_"
    )

    msg2 = (
        f":blue_car: *Vehicles That Moved Today ({len(moved_vehicles)})*\n"
        f"{moving_table}"
    )

    msg3 = (
        f":parking: *Parked Vehicles — Online ({len(stationary)})*\n"
        f"{stationary_table}\n"
        f"_:red_circle: {len(offline_vehicles)} vehicles offline — excluded from list_"
    )

    # Speeding alert message
    if speed_violations:
        spd_table = "```\n"
        spd_table += f"{'Vehicle':<35} {'Max Speed':>10} {'Avg Speed':>10} {'Trip km':>8}\n"
        spd_table += "─" * 68 + "\n"
        for sv in sorted(speed_violations, key=lambda x: x['max_speed'], reverse=True):
            spd_table += f"{sv['vehicle'][:34]:<35} {sv['max_speed']:>8.0f} km {sv['avg_speed']:>8.0f} km {sv['mileage']:>6.1f}km\n"
        spd_table += "```"
        msg4 = (
            f":rotating_light: *Speeding Alert — Yesterday ({len(speed_violations)} events over 120 km/h)*\n"
            f"{spd_table}"
        )
    else:
        msg4 = ":white_check_mark: *No speeding events recorded yesterday*"
    try:
        headers = {"Authorization": f"Bearer {SLACK_TOKEN}", "Content-Type": "application/json"}
        for m in [msg, msg2, msg3, msg4]:
            r = requests.post("https://slack.com/api/chat.postMessage",
                              headers=headers,
                              json={"channel": SLACK_CHANNEL, "text": m}, timeout=10)
            result = r.json()
            if not result.get("ok"):
                print(f"  [Slack] Failed: {result.get('error')}")
        print("  [Slack] Posted 4 messages to #daily-gps-update ✓")
    except Exception as e:
        print(f"  [Slack] Error: {e}")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="Report date YYYY-MM-DD (default: today)")
    args = parser.parse_args()

    report_date = datetime.now(DUBAI_OFFSET)
    if args.date:
        report_date = datetime.strptime(args.date, "%Y-%m-%d").replace(tzinfo=DUBAI_OFFSET)
    date_str = report_date.strftime("%Y-%m-%d")

    print(f"\n{'='*55}")
    print(f"  MKV CAR RENTAL — Daily GPS Report")
    print(f"  Date: {date_str}")
    print(f"{'='*55}\n")

    api = PilotAPI()

    print("Fetching vehicle data...")
    vehicles = api.get_vehicles()
    if not vehicles:
        print("ERROR: No vehicles returned."); sys.exit(1)

    print(f"Fetching yesterday's mileage for {len(vehicles)} vehicles...")
    vehicle_ids = [v["id"] for v in vehicles]
    daily_mileage = api.get_daily_mileage(vehicle_ids, report_date)
    for v in vehicles:
        v["daily_km"] = daily_mileage.get(v["id"], 0)
    total_daily_km = sum(v["daily_km"] for v in vehicles)
    print(f"  Total km yesterday: {total_daily_km:,.1f} km")

    print("Fetching speeding events...")
    speed_violations = api.get_speed_violations(
        [v for v in vehicles if v.get("daily_km", 0) > 0], report_date)
    print(f"  Speeding events (>120 km/h): {len(speed_violations)}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Building Summary...")
    build_summary(wb, date_str, vehicles)
    print("Building Daily Movement...")
    build_daily_movement(wb, vehicles, date_str)
    print("Building Activity Report...")
    build_activity(wb, vehicles, date_str)
    print("Building Speed & Status...")
    build_speed_status(wb, vehicles, date_str)
    print("Building Geofence...")
    build_geofence(wb, vehicles, date_str)

    fname = f"MKV_GPS_Report_{date_str}.xlsx"
    out   = os.path.join(os.path.dirname(os.path.abspath(__file__)), fname)
    wb.save(out)

    online = sum(1 for v in vehicles if v["online"])
    moving = sum(1 for v in vehicles if v["is_moving"])
    moved  = sum(1 for v in vehicles if v["moved_today"])

    print(f"\n✓ Report saved: {fname}")
    print(f"  Vehicles: {len(vehicles)} | Online: {online} | Moving: {moving} | Moved today: {moved}")
    print(f"  Km driven yesterday: {total_daily_km:,.1f} km")

    print("\nPosting to Slack...")
    post_slack(date_str, vehicles, speed_violations)
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
